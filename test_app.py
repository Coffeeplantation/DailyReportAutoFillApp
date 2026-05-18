"""
ウェブ版 全行程自動テスト
対象: app.py (Flask アプリ)
"""
import base64
import io
import pytest
import openpyxl
from datetime import date, time as dtime

from app import app, parse_time, find_column, find_header_row


# ── ヘルパー ─────────────────────────────────────────────────────

def basic_auth(username='root', password='root'):
    creds = base64.b64encode(f'{username}:{password}'.encode()).decode()
    return {'Authorization': f'Basic {creds}'}


def make_xlsx(headers=None):
    """ヘッダー付きの最小限テスト用 xlsx を返す (BytesIO)。"""
    if headers is None:
        headers = {'A': '開始時間', 'B': '終了時間', 'C': '休憩時間', 'D': '備考'}
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, label in headers.items():
        ws[f'{col}1'] = label
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


BASE_TIMES = {
    'monday_start': '09:00', 'monday_end': '17:30', 'monday_break': '01:00',
    'tuesday_start': '09:00', 'tuesday_end': '17:30', 'tuesday_break': '01:00',
    'wednesday_start': '09:00', 'wednesday_end': '17:30', 'wednesday_break': '01:00',
    'thursday_start': '09:00', 'thursday_end': '17:30', 'thursday_break': '01:00',
    'friday_start': '09:00', 'friday_end': '17:30', 'friday_break': '01:00',
    'label_start': '開始時間', 'label_end': '終了時間',
    'label_break': '休憩時間', 'label_note': '備考',
    'note_workday': '在宅勤務', 'paid_leave_dates': '',
}


@pytest.fixture
def client():
    app.config['TESTING'] = True
    with app.test_client() as c:
        yield c


def post_write(client, extra=None, xlsx_buf=None, year='2026', month='5'):
    if xlsx_buf is None:
        xlsx_buf = make_xlsx()
    data = {'year': year, 'month': month,
            'excel_file': (xlsx_buf, 'test.xlsx'),
            **BASE_TIMES}
    if extra:
        data.update(extra)
    return client.post('/api/write', data=data,
                       headers=basic_auth(),
                       content_type='multipart/form-data')


def load_wb(response):
    return openpyxl.load_workbook(io.BytesIO(response.data))


# ── ユニットテスト ────────────────────────────────────────────────

class TestParseTime:
    def test_normal(self):
        t = parse_time('09:00')
        assert t == dtime(9, 0)

    def test_with_seconds(self):
        t = parse_time('17:30:00')
        assert t == dtime(17, 30)

    def test_empty_string(self):
        assert parse_time('') is None

    def test_none(self):
        assert parse_time(None) is None

    def test_invalid(self):
        assert parse_time('abc') is None

    def test_midnight(self):
        assert parse_time('00:00') == dtime(0, 0)


class TestFindColumn:
    def _ws(self, mapping):
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, val in mapping.items():
            ws[f'{col}1'] = val
        return ws

    def test_exact_match(self):
        ws = self._ws({'A': '開始時間', 'B': '終了時間'})
        assert find_column(ws, 1, '開始時間') == 'A'
        assert find_column(ws, 1, '終了時間') == 'B'

    def test_partial_match(self):
        ws = self._ws({'C': '開始時間（HH:MM）'})
        assert find_column(ws, 1, '開始時間') == 'C'

    def test_not_found(self):
        ws = self._ws({'A': '日付'})
        assert find_column(ws, 1, '開始時間') is None

    def test_empty_label(self):
        ws = self._ws({'A': '開始時間'})
        assert find_column(ws, 1, '') is None


class TestFindHeaderRow:
    def test_finds_correct_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A3'] = '開始時間'
        ws['B3'] = '終了時間'
        ws['C3'] = '備考'
        assert find_header_row(ws, ['開始時間', '終了時間', '備考']) == 3

    def test_not_found(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '日付'
        assert find_header_row(ws, ['開始時間', '終了時間']) is None

    def test_minimum_two_labels(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A5'] = '開始時間'
        ws['B5'] = '終了時間'
        # 1つだけでは見つからない
        assert find_header_row(ws, ['開始時間', '存在しない']) is None
        # 2つあれば見つかる
        assert find_header_row(ws, ['開始時間', '終了時間']) == 5


# ── 認証テスト ────────────────────────────────────────────────────

class TestAuth:
    def test_no_credentials_index(self, client):
        assert client.get('/').status_code == 401

    def test_wrong_password(self, client):
        assert client.get('/', headers=basic_auth('root', 'wrong')).status_code == 401

    def test_correct_credentials(self, client):
        assert client.get('/', headers=basic_auth()).status_code == 200

    def test_no_credentials_holidays(self, client):
        assert client.get('/api/holidays?year=2026&month=1').status_code == 401

    def test_no_credentials_write(self, client):
        assert client.post('/api/write', data={}).status_code == 401

    def test_no_credentials_manual(self, client):
        assert client.get('/api/manual').status_code == 401

    def test_no_credentials_report(self, client):
        assert client.get('/api/report').status_code == 401


# ── メインページ ──────────────────────────────────────────────────

class TestIndex:
    def test_status_200(self, client):
        r = client.get('/', headers=basic_auth())
        assert r.status_code == 200

    def test_contains_form_fields(self, client):
        html = client.get('/', headers=basic_auth()).data.decode()
        for field in ['monday_start', 'monday_end', 'monday_break',
                      'paid_leave_dates', 'excel_file', 'label_note']:
            assert field in html, f'フィールドが見つからない: {field}'

    def test_contains_current_year(self, client):
        html = client.get('/', headers=basic_auth()).data.decode()
        assert str(date.today().year) in html

    def test_contains_current_month(self, client):
        html = client.get('/', headers=basic_auth()).data.decode()
        assert str(date.today().month) in html


# ── 祝日 API ─────────────────────────────────────────────────────

class TestHolidaysAPI:
    def test_new_year_day(self, client):
        r = client.get('/api/holidays?year=2026&month=1', headers=basic_auth())
        assert r.status_code == 200
        data = r.get_json()
        assert '1' in data          # 元日が含まれる
        assert '元日' in data['1']  # 名称確認

    def test_golden_week(self, client):
        r = client.get('/api/holidays?year=2026&month=5', headers=basic_auth())
        data = r.get_json()
        assert '3' in data   # 憲法記念日
        assert '4' in data   # みどりの日
        assert '5' in data   # こどもの日

    def test_month_with_no_holidays(self, client):
        # 祝日のない月 (例: 6月)
        r = client.get('/api/holidays?year=2026&month=6', headers=basic_auth())
        assert r.status_code == 200
        assert r.get_json() == {}

    def test_invalid_month_13(self, client):
        r = client.get('/api/holidays?year=2026&month=13', headers=basic_auth())
        assert r.status_code == 400

    def test_invalid_month_0(self, client):
        r = client.get('/api/holidays?year=2026&month=0', headers=basic_auth())
        assert r.status_code == 400

    def test_invalid_year_too_old(self, client):
        r = client.get('/api/holidays?year=1999&month=1', headers=basic_auth())
        assert r.status_code == 400

    def test_invalid_year_too_future(self, client):
        r = client.get('/api/holidays?year=2100&month=1', headers=basic_auth())
        assert r.status_code == 400

    def test_invalid_non_numeric(self, client):
        r = client.get('/api/holidays?year=abc&month=1', headers=basic_auth())
        assert r.status_code == 400


# ── 書き込み API ─────────────────────────────────────────────────

class TestWriteAPI:

    # ── エラー系 ──

    def test_no_file(self, client):
        r = client.post('/api/write',
                        data={'year': '2026', 'month': '5'},
                        headers=basic_auth(),
                        content_type='multipart/form-data')
        assert r.status_code == 400

    def test_invalid_file_format(self, client):
        data = {'excel_file': (io.BytesIO(b'not xlsx'), 'fake.xlsx'),
                'year': '2026', 'month': '5'}
        r = client.post('/api/write', data=data, headers=basic_auth(),
                        content_type='multipart/form-data')
        assert r.status_code == 400

    def test_invalid_year(self, client):
        r = post_write(client, year='1999')
        assert r.status_code == 400

    def test_invalid_month(self, client):
        r = post_write(client, month='13')
        assert r.status_code == 400

    def test_invalid_year_too_future(self, client):
        r = post_write(client, year='2100')
        assert r.status_code == 400

    # ── 正常系：レスポンス確認 ──

    def test_returns_xlsx(self, client):
        r = post_write(client)
        assert r.status_code == 200
        assert 'spreadsheetml' in r.content_type

    def test_response_is_valid_xlsx(self, client):
        r = post_write(client)
        wb = load_wb(r)
        assert wb.active is not None

    # ── 出勤日 ──
    # 2026-05-01 = 金曜日 → row 2

    def test_workday_friday_start(self, client):
        r = post_write(client, year='2026', month='5')
        ws = load_wb(r).active
        # A2 = 開始時間 → time(9,0)
        assert ws['A2'].value == dtime(9, 0)

    def test_workday_friday_end(self, client):
        r = post_write(client, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['B2'].value == dtime(17, 30)

    def test_workday_friday_break(self, client):
        r = post_write(client, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['C2'].value == dtime(1, 0)

    def test_workday_note(self, client):
        r = post_write(client, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['D2'].value == '在宅勤務'

    # ── 土日 ──
    # 2026-05-02 = 土曜 → row 3、2026-05-03 = 日曜 (かつ祝日) → row 4

    def test_saturday_is_blank(self, client):
        r = post_write(client, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['A3'].value is None
        assert ws['D3'].value is None

    # ── 祝日 ──
    # 2026-01-01 = 元日（木曜） → row 2

    def test_holiday_note(self, client):
        r = post_write(client, year='2026', month='1')
        ws = load_wb(r).active
        assert ws['D2'].value == '祝日'

    def test_holiday_start_is_blank(self, client):
        r = post_write(client, year='2026', month='1')
        ws = load_wb(r).active
        assert ws['A2'].value is None

    # ── 有給取得日 ──
    # 2026-05-07 = 木曜 → row 8

    def test_paid_leave_note(self, client):
        r = post_write(client, extra={'paid_leave_dates': '7'},
                       year='2026', month='5')
        ws = load_wb(r).active
        assert ws['D8'].value == '私用により、休暇'

    def test_paid_leave_start_is_blank(self, client):
        r = post_write(client, extra={'paid_leave_dates': '7'},
                       year='2026', month='5')
        ws = load_wb(r).active
        assert ws['A8'].value is None

    def test_multiple_paid_leave(self, client):
        # 7日と8日（木・金）を有給
        r = post_write(client, extra={'paid_leave_dates': '7,8'},
                       year='2026', month='5')
        ws = load_wb(r).active
        assert ws['D8'].value == '私用により、休暇'
        assert ws['D9'].value == '私用により、休暇'

    # ── 時間例外日 ──
    # 2026-05-01 (金) を残業: 10:00-21:00

    def test_time_exception_start(self, client):
        extra = {'time_ex_day': '1', 'time_ex_start': '10:00',
                 'time_ex_end': '21:00', 'time_ex_break': '01:00'}
        r = post_write(client, extra=extra, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['A2'].value == dtime(10, 0)

    def test_time_exception_end(self, client):
        extra = {'time_ex_day': '1', 'time_ex_start': '10:00',
                 'time_ex_end': '21:00', 'time_ex_break': '01:00'}
        r = post_write(client, extra=extra, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['B2'].value == dtime(21, 0)

    def test_time_exception_overrides_weekend(self, client):
        # 土曜日 (2026-05-02) に時間例外 → 出勤扱いになる
        extra = {'time_ex_day': '2', 'time_ex_start': '09:00',
                 'time_ex_end': '17:30', 'time_ex_break': '01:00'}
        r = post_write(client, extra=extra, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['A3'].value == dtime(9, 0)

    def test_time_exception_overrides_holiday(self, client):
        # 元日 (2026-01-01) に時間例外 → 出勤扱い
        extra = {'time_ex_day': '1', 'time_ex_start': '09:00',
                 'time_ex_end': '17:30', 'time_ex_break': '01:00'}
        r = post_write(client, extra=extra, year='2026', month='1')
        ws = load_wb(r).active
        assert ws['A2'].value == dtime(9, 0)

    # ── 備考例外 ──

    def test_note_exception(self, client):
        # 2026-05-13 (水) の備考を個別設定（祝日でない平日）
        extra = {'exception_day': '13', 'exception_note': '客先常駐'}
        r = post_write(client, extra=extra, year='2026', month='5')
        ws = load_wb(r).active
        assert ws['D14'].value == '客先常駐'

    # ── カスタムラベル ──

    def test_custom_label_detection(self, client):
        # ヘッダーを日本語以外のラベルにしても列を検出できる
        xlsx = make_xlsx({'A': 'Start', 'B': 'End', 'C': 'Break', 'D': 'Note'})
        extra = {'label_start': 'Start', 'label_end': 'End',
                 'label_break': 'Break', 'label_note': 'Note'}
        r = post_write(client, extra=extra, xlsx_buf=xlsx,
                       year='2026', month='5')
        ws = load_wb(r).active
        assert ws['D2'].value == '在宅勤務'

    def test_custom_note_workday(self, client):
        r = post_write(client, extra={'note_workday': '出社勤務'},
                       year='2026', month='5')
        ws = load_wb(r).active
        # 2026-05-01 (金) の備考
        assert ws['D2'].value == '出社勤務'

    # ── 曜日別の時間が正しく適用されるか ──

    def test_each_weekday_start_time(self, client):
        # 月〜金それぞれ異なる開始時間
        extra = {
            'monday_start': '08:00',
            'tuesday_start': '08:30',
            'wednesday_start': '09:00',
            'thursday_start': '09:30',
            'friday_start': '10:00',
        }
        r = post_write(client, extra=extra, year='2026', month='5')
        ws = load_wb(r).active
        # 2026-05: 1=金, 7=木, 8=金, 11=月, 12=火, 13=水, 14=木, 15=金
        assert ws['A12'].value == dtime(8, 0)   # 月曜 (11日)
        assert ws['A13'].value == dtime(8, 30)  # 火曜 (12日)
        assert ws['A14'].value == dtime(9, 0)   # 水曜 (13日)
        assert ws['A15'].value == dtime(9, 30)  # 木曜 (14日)
        assert ws['A16'].value == dtime(10, 0)  # 金曜 (15日)

    # ── ヘッダー行の自動検出 ──

    def test_header_row_detection(self, client):
        # ヘッダーが先頭行ではなく途中の行にある場合
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A5'] = '開始時間'
        ws['B5'] = '終了時間'
        ws['C5'] = '休憩時間'
        ws['D5'] = '備考'
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        r = post_write(client, xlsx_buf=buf, year='2026', month='5')
        assert r.status_code == 200
        ws_out = load_wb(r).active
        # Day 1 = row 6 (header=5, data_start=6)
        assert ws_out['D6'].value == '在宅勤務'

    # ── カーソル位置（A1）──

    def test_cursor_at_a1(self, client):
        r = post_write(client, year='2026', month='5')
        wb = load_wb(r)
        ws = wb.active
        sel = ws.sheet_view.selection[0]
        assert sel.activeCell == 'A1'


# ── マニュアル・レポートダウンロード ────────────────────────────

class TestDownloads:
    def test_manual_status(self, client):
        r = client.get('/api/manual', headers=basic_auth())
        assert r.status_code == 200

    def test_manual_content_type(self, client):
        r = client.get('/api/manual', headers=basic_auth())
        assert 'spreadsheetml' in r.content_type

    def test_manual_is_valid_xlsx(self, client):
        r = client.get('/api/manual', headers=basic_auth())
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        assert '使い方ガイド' in wb.sheetnames

    def test_report_status(self, client):
        r = client.get('/api/report', headers=basic_auth())
        assert r.status_code == 200

    def test_report_content_type(self, client):
        r = client.get('/api/report', headers=basic_auth())
        assert 'spreadsheetml' in r.content_type

    def test_report_is_valid_xlsx(self, client):
        r = client.get('/api/report', headers=basic_auth())
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        assert '表紙' in wb.sheetnames

    def test_manual_no_auth(self, client):
        assert client.get('/api/manual').status_code == 401

    def test_report_no_auth(self, client):
        assert client.get('/api/report').status_code == 401
