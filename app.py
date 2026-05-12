
from flask import Flask, render_template, request, send_file, jsonify
from flask_httpauth import HTTPBasicAuth
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import jpholiday
from datetime import date, time
import calendar
import io
import os
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 4 * 1024 * 1024  # 4MB (Vercel制限に合わせる)
auth = HTTPBasicAuth()

USERS = {
    os.getenv('APP_USERNAME', 'root'): generate_password_hash(os.getenv('APP_PASSWORD', 'root')),
}

@auth.verify_password
def verify_password(username, password):
    if username in USERS and check_password_hash(USERS[username], password):
        return username


def parse_time(time_str):
    if not time_str:
        return None
    try:
        parts = time_str.split(':')  # HH:MM または HH:MM:SS に対応
        return time(int(parts[0]), int(parts[1]))
    except Exception:
        return None


def find_column(ws, header_row, label):
    """ヘッダー行からラベルに一致する列を返す（部分一致対応）。見つからなければ None。"""
    label = label.strip()
    if not label:
        return None
    for cell in ws[header_row]:
        if cell.value:
            cell_val = str(cell.value).strip()
            if cell_val == label or label in cell_val or cell_val in label:
                return cell.column_letter
    return None


def find_header_row(ws, labels):
    """ラベルを2つ以上含む行をヘッダー行として検索する。見つからなければ None。"""
    clean = [l.strip() for l in labels if l.strip()]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_vals = [str(c.value).strip() for c in row if c.value]
        hits = sum(
            1 for label in clean
            if any(label in v or v in label for v in row_vals)
        )
        if hits >= 2:
            return row[0].row
    return None


@app.route('/')
@auth.login_required
def index():
    today = date.today()
    return render_template('index.html',
                           current_month=today.month,
                           current_year=today.year)


@app.route('/api/holidays')
@auth.login_required
def get_holidays():
    today = date.today()
    try:
        year  = int(request.args.get('year',  today.year))
        month = int(request.args.get('month', today.month))
        if not (2000 <= year <= 2099) or not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({'error': '年月の値が正しくありません'}), 400
    _, last_day = calendar.monthrange(year, month)
    holidays = {}
    for day in range(1, last_day + 1):
        name = jpholiday.is_holiday_name(date(year, month, day))
        if name:
            holidays[str(day)] = name
    return jsonify(holidays)


@app.route('/api/write', methods=['POST'])
@auth.login_required
def write_excel():
    excel_file = request.files.get('excel_file')
    if not excel_file:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    safe_name = secure_filename(excel_file.filename) or 'report.xlsx'

    # 有給取得日
    paid_leave_str = request.form.get('paid_leave_dates', '')
    paid_leave = set()
    for d in paid_leave_str.split(','):
        d = d.strip()
        if d.isdigit():
            paid_leave.add(int(d))

    # 曜日別勤務時間（0=月〜4=金）
    day_names = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
    weekday_times = {}
    for i, name in enumerate(day_names):
        start = parse_time(request.form.get(f'{name}_start'))
        end = parse_time(request.form.get(f'{name}_end'))
        brk = parse_time(request.form.get(f'{name}_break')) or time(1, 0)
        if start and end:
            weekday_times[i] = {'start': start, 'end': end, 'break': brk}

    today = date.today()
    try:
        year  = int(request.form.get('year',  today.year))
        month = int(request.form.get('month', today.month))
        if not (2000 <= year <= 2099) or not (1 <= month <= 12):
            raise ValueError
    except (ValueError, TypeError):
        return jsonify({'error': '年月の値が正しくありません（年：2000〜2099、月：1〜12）'}), 400
    _, last_day = calendar.monthrange(year, month)

    # 祝日取得
    holidays = set()
    for day in range(1, last_day + 1):
        if jpholiday.is_holiday(date(year, month, day)):
            holidays.add(day)

    # Excel処理
    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception:
        return jsonify({'error': 'ファイルを開けませんでした。Excel形式（.xlsx）のファイルを選択してください。'}), 400
    ws = wb.active

    label_start  = request.form.get('label_start',  '開始時間')
    label_end    = request.form.get('label_end',    '終了時間')
    label_break  = request.form.get('label_break',  '休憩時間')
    label_note   = request.form.get('label_note',   '備考')
    note_workday = request.form.get('note_workday', '在宅勤務')

    # ヘッダー行をラベルで動的検索（見つからなければ19行目をフォールバック）
    HEADER_ROW    = find_header_row(ws, [label_start, label_end, label_break, label_note]) or 19
    DATA_START_ROW = HEADER_ROW + 1

    # 例外日の備考（日付→備考のマッピング）
    ex_days  = request.form.getlist('exception_day')
    ex_notes = request.form.getlist('exception_note')
    note_exceptions = {}
    for d, n in zip(ex_days, ex_notes):
        if d.isdigit():
            note_exceptions[int(d)] = n.strip() or None

    # 例外日の勤務時間（日付→時間のマッピング）
    time_ex_days   = request.form.getlist('time_ex_day')
    time_ex_starts = request.form.getlist('time_ex_start')
    time_ex_ends   = request.form.getlist('time_ex_end')
    time_ex_breaks = request.form.getlist('time_ex_break')
    time_exceptions = {}
    for d, s, e, b in zip(time_ex_days, time_ex_starts, time_ex_ends, time_ex_breaks):
        if d.isdigit():
            start = parse_time(s)
            end   = parse_time(e)
            brk   = parse_time(b) or time(1, 0)
            if start and end:
                time_exceptions[int(d)] = {'start': start, 'end': end, 'break': brk}

    col_start = find_column(ws, HEADER_ROW, label_start) or 'F'
    col_end   = find_column(ws, HEADER_ROW, label_end)   or 'I'
    col_break = find_column(ws, HEADER_ROW, label_break) or 'L'
    col_note  = find_column(ws, HEADER_ROW, label_note)  or 'S'

    black_font = Font(color='000000')

    for day in range(1, last_day + 1):
        row = DATA_START_ROW + (day - 1)
        weekday = date(year, month, day).weekday()  # 0=月, 6=日

        # 対象セルをクリア
        for col in [col_start, col_end, col_break, col_note]:
            ws[f'{col}{row}'].value = None

        if day in time_exceptions:
            # 時間例外が最優先（土日・祝日・有給より上書き可能）
            t = time_exceptions[day]
            for col, val, fmt in [
                (col_start, t['start'], 'h:mm'),
                (col_end,   t['end'],   'h:mm'),
                (col_break, t['break'], 'h:mm'),
            ]:
                ws[f'{col}{row}'].value = val
                ws[f'{col}{row}'].number_format = fmt
                ws[f'{col}{row}'].font = black_font
            ws[f'{col_note}{row}'].value = note_exceptions[day] if day in note_exceptions else note_workday
            ws[f'{col_note}{row}'].font = black_font
        elif day in paid_leave:
            ws[f'{col_note}{row}'].value = '私用により、休暇'
            ws[f'{col_note}{row}'].font = black_font
        elif weekday >= 5:
            pass  # 土日：空欄
        elif day in holidays:
            ws[f'{col_note}{row}'].value = '祝日'
            ws[f'{col_note}{row}'].font = black_font
        elif weekday in weekday_times:
            t = weekday_times[weekday]
            for col, val, fmt in [
                (col_start, t['start'], 'h:mm'),
                (col_end,   t['end'],   'h:mm'),
                (col_break, t['break'], 'h:mm'),
            ]:
                ws[f'{col}{row}'].value = val
                ws[f'{col}{row}'].number_format = fmt
                ws[f'{col}{row}'].font = black_font
            ws[f'{col_note}{row}'].value = note_exceptions[day] if day in note_exceptions else note_workday
            ws[f'{col_note}{row}'].font = black_font

    # カーソルをA1に設定
    try:
        ws.sheet_view.selection[0].activeCell = 'A1'
        ws.sheet_view.selection[0].sqref = 'A1'
    except (IndexError, AttributeError):
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=safe_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _build_manual() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "使い方ガイド"

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 52
    ws.column_dimensions['D'].width = 28

    C_NAVY  = "1A56DB"; C_ACCENT = "DBEAFE"; C_GREEN = "059669"
    C_ORANGE = "EA580C"; C_GRAY = "F3F4F6"; C_WHITE = "FFFFFF"; C_BORDER = "CBD5E1"

    def _side(c=C_BORDER): return Side(border_style="thin", color=c)
    def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    def _fill(c): return PatternFill("solid", fgColor=c)
    def _font(bold=False, size=11, color="000000", italic=False):
        return Font(name="Meiryo", bold=bold, size=size, color=color, italic=italic)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _merge(r1, c1, r2, c2): ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    def _rh(row, h): ws.row_dimensions[row].height = h

    # ── タイトル ──
    for r in range(1, 5): _rh(r, 6 if r in (1, 4) else 24)
    for r in range(1, 5):
        for c in range(1, 5): ws.cell(r, c).fill = _fill(C_NAVY)
    _merge(2, 1, 3, 4)
    c = ws.cell(2, 1, "日報自動入力アプリ  使い方ガイド")
    c.font = Font(name="Meiryo", bold=True, size=18, color=C_WHITE)
    c.alignment = _align("center", "center")

    # ── アプリ概要 ──
    _rh(5, 8)
    _merge(6, 1, 6, 4); _rh(6, 28)
    c = ws.cell(6, 1, "■  アプリ概要")
    c.font = _font(bold=True, size=13, color=C_WHITE); c.fill = _fill(C_NAVY)
    c.alignment = _align(); c.border = _border()

    overview = [
        ("目的",     "毎月会社から配布されるExcel作業報告書に、勤務時間・備考を自動で書き込み・保存するWebアプリです。"),
        ("メリット", "月初に最小限の入力をするだけで、1ヶ月分のExcelが自動で完成します。手入力の手間をゼロにします。"),
        ("対象",     "Excel形式（.xlsx）の作業報告書を毎月提出している社員が対象です。"),
        ("動作環境", "ブラウザ（Chrome / Edge / Safari など）で動作します。インストール不要です。"),
    ]
    for i, (label, text) in enumerate(overview):
        r = 7 + i; _rh(r, 36)
        c1 = ws.cell(r, 2, label)
        c1.font = _font(bold=True, size=10, color=C_NAVY); c1.fill = _fill(C_ACCENT)
        c1.alignment = _align("center"); c1.border = _border()
        _merge(r, 3, r, 4)
        c2 = ws.cell(r, 3, text)
        c2.font = _font(size=10); c2.fill = _fill(C_WHITE)
        c2.alignment = _align(wrap=True); c2.border = _border()

    # ── 使い方手順 ──
    _rh(12, 8)
    _merge(13, 1, 13, 4); _rh(13, 28)
    c = ws.cell(13, 1, "■  使い方手順")
    c.font = _font(bold=True, size=13, color=C_WHITE); c.fill = _fill(C_GREEN)
    c.alignment = _align(); c.border = _border()

    steps = [
        ("STEP 1", "アプリを開く",
         "ブラウザでアプリのURLにアクセスします。ユーザー名・パスワードを入力してログインしてください。"),
        ("STEP 2", "対象月を確認する",
         "画面上部に「年・月」が表示されます。自動で当月が設定されます。\n異なる月を処理したい場合は数字を直接変更してください。"),
        ("STEP 3", "曜日別の勤務時間を入力する",
         "月〜金それぞれの開始・終了・休憩時間を入力します。\n設定はブラウザに自動保存され、次回起動時に引き継がれます。"),
        ("STEP 4", "例外日を設定する（任意）",
         "「＋ 例外日を追加」から残業・早退・休日出勤など通常と異なる日を個別設定できます。\n土日・祝日も選択可能で、例外日設定はすべての設定より優先されます。"),
        ("STEP 5", "有給取得日を選択する（任意）",
         "「有給を取得する日がある」にチェックを入れるとカレンダーが表示されます。\n取得する日をクリックして選択してください（土日・祝日も選択可能）。"),
        ("STEP 6", "Excelファイルを選択する",
         "ファイル選択エリアをクリックし、会社から配布された作業報告書（.xlsx）を選びます。\n選択後、✕ ボタンで選択をやり直せます。"),
        ("STEP 7", "入力完了・ダウンロード",
         "「入力完了・ダウンロード」ボタンを押すと、自動入力済みのExcelファイルがダウンロードされます。\nファイルを開いて内容を確認し、所定の場所に保存してください。"),
    ]
    step_colors = [
        ("1E40AF","DBEAFE"),("065F46","D1FAE5"),("92400E","FEF3C7"),
        ("7C3AED","EDE9FE"),("BE185D","FCE7F3"),("1E40AF","DBEAFE"),("065F46","D1FAE5"),
    ]
    cur = 14
    for i, (step, title, desc) in enumerate(steps):
        fc, bc = step_colors[i]; _rh(cur, 14)
        c1 = ws.cell(cur, 2, step)
        c1.font = Font(name="Meiryo", bold=True, size=9, color=C_WHITE)
        c1.fill = _fill(fc); c1.alignment = _align("center"); c1.border = _border()
        ws.merge_cells(start_row=cur, start_column=3, end_row=cur, end_column=4)
        c2 = ws.cell(cur, 3, title)
        c2.font = Font(name="Meiryo", bold=True, size=11, color=fc)
        c2.fill = _fill(bc); c2.alignment = _align(); c2.border = _border()
        ws.cell(cur, 4).fill = _fill(bc); ws.cell(cur, 4).border = _border()
        cur += 1
        lc = desc.count('\n') + 1; _rh(cur, max(32, lc * 28))
        ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=4)
        c3 = ws.cell(cur, 2, desc)
        c3.font = _font(size=10); c3.fill = _fill(C_WHITE)
        c3.alignment = _align(wrap=True); c3.border = _border()
        cur += 1; _rh(cur, 5); cur += 1

    # ── 自動入力ルール ──
    _rh(cur, 8); cur += 1
    _merge(cur, 1, cur, 4); _rh(cur, 28)
    c = ws.cell(cur, 1, "■  自動入力ルール")
    c.font = _font(bold=True, size=13, color=C_WHITE); c.fill = _fill(C_ORANGE)
    c.alignment = _align(); c.border = _border(); cur += 1
    _rh(cur, 22)
    for j, h in enumerate(["日の種類", "開始・終了・休憩", "備考欄"]):
        cx = ws.cell(cur, 2 + j, h)
        cx.font = _font(bold=True, size=10, color=C_WHITE); cx.fill = _fill(C_ORANGE)
        cx.alignment = _align("center"); cx.border = _border()
    cur += 1
    for kind, times, note, bg in [
        ("出勤日（平日）",    "曜日別設定の時間・休憩 1:00", "在宅勤務（変更可）", C_WHITE),
        ("例外日（休日出勤）","例外設定した時間・休憩 1:00", "在宅勤務（変更可）", "FFF9C4"),
        ("有給取得日",        "空欄",                        "私用により、休暇",   "FCE7F3"),
        ("祝日",              "空欄",                        "祝日",               "FFF5F5"),
        ("土日",              "空欄",                        "空欄",               C_GRAY),
    ]:
        _rh(cur, 28)
        for col, val in zip([2, 3, 4], [kind, times, note]):
            cx = ws.cell(cur, col, val)
            cx.font = _font(size=10); cx.fill = _fill(bg)
            cx.alignment = _align("center"); cx.border = _border()
        cur += 1

    # ── よくある疑問 ──
    _rh(cur, 8); cur += 1
    _merge(cur, 1, cur, 4); _rh(cur, 28)
    c = ws.cell(cur, 1, "■  よくある疑問")
    c.font = _font(bold=True, size=13, color=C_WHITE); c.fill = _fill("6D28D9")
    c.alignment = _align(); c.border = _border(); cur += 1
    for q, a in [
        ("設定は毎月入力し直す？",
         "いいえ。曜日別の勤務時間・各種ラベルはブラウザに自動保存されます。次回起動時には前回の設定が自動で反映されます。"),
        ("休日出勤はどう入力する？",
         "「＋ 例外日を追加」から該当日を選択（土日・祝日も選択可）し、出勤時間を入力します。例外日設定は最優先で反映されます。"),
        ("有給を祝日や土日に取得したい場合は？",
         "有給カレンダーで土日・祝日もクリック選択できます。選択した日は「私用により、休暇」が記入されます。"),
        ("ファイルを間違えて選択した場合は？",
         "ファイル名の右に表示される「✕」ボタンを押すと選択をキャンセルできます。再度クリックして正しいファイルを選び直してください。"),
        ("書き込まれる列や行がずれている場合は？",
         "アプリが開始時間・終了時間・休憩時間・備考のヘッダーを自動検索して列を特定します。Excelのヘッダー文字列と設定が一致しているか確認してください。"),
    ]:
        _rh(cur, 18)
        _merge(cur, 2, cur, 4)
        cq = ws.cell(cur, 2, f"Q.  {q}")
        cq.font = Font(name="Meiryo", bold=True, size=10, color="6D28D9")
        cq.fill = _fill("EDE9FE"); cq.alignment = _align(); cq.border = _border(); cur += 1
        _rh(cur, max(32, a.count('\n') * 28 + 28))
        _merge(cur, 2, cur, 4)
        ca = ws.cell(cur, 2, f"A.  {a}")
        ca.font = _font(size=10); ca.fill = _fill(C_WHITE)
        ca.alignment = _align(wrap=True); ca.border = _border(); cur += 1
        _rh(cur, 4); cur += 1

    # ── フッター ──
    _rh(cur, 8); cur += 1
    _merge(cur, 1, cur, 4); _rh(cur, 22)
    cf = ws.cell(cur, 1, "日報自動入力アプリ  操作マニュアル  ／  本ファイルはアプリから自動生成されました")
    cf.font = _font(size=9, color="94A3B8", italic=True); cf.fill = _fill(C_NAVY)
    cf.alignment = _align("center")
    for r in range(1, cur + 1): ws.cell(r, 1).fill = _fill(C_NAVY)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


@app.route('/api/manual')
@auth.login_required
def download_manual():
    buf = _build_manual()
    return send_file(
        buf,
        as_attachment=True,
        download_name='日報アプリ_使い方ガイド.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _build_report() -> io.BytesIO:
    from datetime import date as _date
    TODAY = _date.today().strftime("%Y年%m月%d日")

    wb = openpyxl.Workbook()

    def _side(c="CBD5E1"): return Side(border_style="thin", color=c)
    def _border(color="CBD5E1"):
        s = _side(color)
        return Border(left=s, right=s, top=s, bottom=s)
    def _fill(c): return PatternFill("solid", fgColor=c)
    def _font(bold=False, size=11, color="1a1a1a", italic=False):
        return Font(name="Meiryo", bold=bold, size=size, color=color, italic=italic)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _rh(ws, row, h): ws.row_dimensions[row].height = h
    def _mg(ws, r1, c1, r2, c2): ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    def _cw(ws, widths):
        for col, w in zip("ABCDEFGH", widths): ws.column_dimensions[col].width = w
    def _sp(ws, row, h=6): _rh(ws, row, h)

    def sec_head(ws, row, text, bg="1A56DB", cols=5):
        _mg(ws, row, 1, row, cols); _rh(ws, row, 26)
        c = ws.cell(row, 1, text)
        c.font = _font(bold=True, size=12, color="FFFFFF")
        c.fill = _fill(bg); c.alignment = _align("left", "center"); c.border = _border(bg)

    # ── 表紙 ──
    ws1 = wb.active; ws1.title = "表紙"
    _cw(ws1, [4, 20, 44, 20, 4]); ws1.sheet_view.showGridLines = False
    for r in range(1, 32):
        _rh(ws1, r, 18)
        for c in range(1, 6): ws1.cell(r, c).fill = _fill("F0F4FF")
    for r in range(5, 17):
        for c in range(2, 5): ws1.cell(r, c).fill = _fill("1A56DB")
    _mg(ws1, 6, 2, 8, 4)
    c = ws1.cell(6, 2, "日報自動入力アプリ")
    c.font = Font(name="Meiryo", bold=True, size=22, color="FFFFFF"); c.alignment = _align("center", "center")
    _mg(ws1, 9, 2, 10, 4)
    c = ws1.cell(9, 2, "導入説明資料")
    c.font = Font(name="Meiryo", bold=True, size=16, color="BFD7FF"); c.alignment = _align("center", "center")
    _mg(ws1, 12, 2, 13, 4)
    c = ws1.cell(12, 2, "― 業務効率化に向けた月次作業報告書の自動化 ―")
    c.font = Font(name="Meiryo", size=10, color="FFFFFF", italic=True); c.alignment = _align("center", "center")
    for i, (label, val) in enumerate([("作成日", TODAY), ("作成部署", ""), ("作成者", ""), ("提出先", "")]):
        r = 19 + i * 2; _rh(ws1, r, 20)
        lc = ws1.cell(r, 2, label)
        lc.font = _font(bold=True, size=10, color="1A56DB"); lc.fill = _fill("EFF6FF")
        lc.alignment = _align("center"); lc.border = _border("BFD7FF")
        _mg(ws1, r, 3, r, 4)
        vc = ws1.cell(r, 3, val)
        vc.font = _font(size=10); vc.fill = _fill("FFFFFF")
        vc.alignment = _align(); vc.border = _border("BFD7FF")
    _mg(ws1, 30, 2, 30, 4)
    c = ws1.cell(30, 2, "本資料は社内業務効率化ツールの導入説明を目的として作成されました。")
    c.font = _font(size=8, color="94A3B8", italic=True); c.alignment = _align("center")

    # ── アプリ概要 ──
    ws2 = wb.create_sheet("1. アプリ概要")
    _cw(ws2, [2, 18, 50, 20, 2]); ws2.sheet_view.showGridLines = False
    r = 1; _sp(ws2, r); r += 1
    _mg(ws2, r, 1, r, 5); _rh(ws2, r, 30)
    c = ws2.cell(r, 1, "1.  アプリ概要")
    c.font = _font(bold=True, size=16, color="1A56DB"); c.alignment = _align("left", "center"); r += 1
    _mg(ws2, r, 1, r, 5); _rh(ws2, r, 20)
    c = ws2.cell(r, 1, "日報自動入力アプリ  導入説明資料")
    c.font = _font(size=9, color="94A3B8", italic=True); c.alignment = _align("left", "center"); r += 1
    _sp(ws2, r, 10); r += 1
    sec_head(ws2, r, "1-1　導入目的", "1E40AF"); r += 1
    for txt in [
        "本アプリは、毎月会社から配布されるExcel形式の作業報告書への勤務時間・備考欄の記入作業を自動化するWebアプリケーションです。",
        "現状、社員は毎月1ヶ月分の勤務時間（開始・終了・休憩）および備考をExcelへ手作業で入力する業務が発生しています。",
        "本ツールを導入することで、月初に最低限の設定を行うだけで1ヶ月分のExcelへの自動入力・保存が完結し、作業工数の大幅な削減が見込まれます。",
    ]:
        _rh(ws2, r, 36); _mg(ws2, r, 2, r, 4)
        c = ws2.cell(r, 2, txt)
        c.font = _font(size=10); c.fill = _fill("F8FAFF")
        c.alignment = _align(wrap=True); c.border = _border("DBEAFE"); r += 1
    _sp(ws2, r, 8); r += 1
    sec_head(ws2, r, "1-2　期待される効果", "059669"); r += 1
    for label, text in [
        ("工数削減",       "月次の手入力作業（約15〜30分/月）をほぼゼロに削減。年間換算で数時間分の業務時間を創出します。"),
        ("入力ミスの防止", "手動入力に伴う時間の誤入力・記入漏れを排除し、正確な作業報告書を自動生成します。"),
        ("操作の簡便性",   "ブラウザのみで動作し、インストール不要。プログラミング知識がなくても月初に数分で操作が完了します。"),
        ("設定の引き継ぎ", "曜日別の勤務時間設定がブラウザに自動保存されるため、翌月以降は確認・変更のみで対応できます。"),
    ]:
        _rh(ws2, r, 38)
        lc = ws2.cell(r, 2, label)
        lc.font = _font(bold=True, size=10, color="065F46"); lc.fill = _fill("D1FAE5")
        lc.alignment = _align("center"); lc.border = _border("6EE7B7")
        _mg(ws2, r, 3, r, 4)
        tc = ws2.cell(r, 3, text)
        tc.font = _font(size=10); tc.fill = _fill("F0FDF4")
        tc.alignment = _align(wrap=True); tc.border = _border("6EE7B7"); r += 1
    _sp(ws2, r, 8); r += 1
    sec_head(ws2, r, "1-3　対象・動作環境", "7C3AED"); r += 1
    for label, text in [
        ("対象ユーザー", "Excel作業報告書（.xlsx形式）を毎月提出している社員"),
        ("対象ファイル", "会社から毎月配布されるExcel形式の作業報告書（.xlsx）"),
        ("動作環境",     "ブラウザ（Chrome / Edge / Safari など）。インストール・設定作業不要"),
        ("アクセス方法", "社内ネットワーク上のURL、またはローカル環境でブラウザから直接アクセス"),
        ("認証",         "ユーザー名・パスワードによるBasic認証（不正アクセス防止）"),
    ]:
        _rh(ws2, r, 28)
        lc = ws2.cell(r, 2, label)
        lc.font = _font(bold=True, size=10, color="5B21B6"); lc.fill = _fill("EDE9FE")
        lc.alignment = _align("center"); lc.border = _border("C4B5FD")
        _mg(ws2, r, 3, r, 4)
        tc = ws2.cell(r, 3, text)
        tc.font = _font(size=10); tc.fill = _fill("FAFAFA")
        tc.alignment = _align(wrap=True); tc.border = _border("C4B5FD"); r += 1

    # ── 操作手順 ──
    ws3 = wb.create_sheet("2. 操作手順")
    _cw(ws3, [2, 14, 52, 20, 2]); ws3.sheet_view.showGridLines = False
    r = 1; _sp(ws3, r); r += 1
    _mg(ws3, r, 1, r, 5); _rh(ws3, r, 30)
    c = ws3.cell(r, 1, "2.  操作手順")
    c.font = _font(bold=True, size=16, color="1A56DB"); c.alignment = _align("left", "center"); r += 1
    _mg(ws3, r, 1, r, 5); _rh(ws3, r, 18)
    c = ws3.cell(r, 1, "月初の操作は以下7ステップで完了します。所要時間の目安：5分以内")
    c.font = _font(size=9, color="94A3B8", italic=True); c.alignment = _align("left", "center"); r += 1
    _sp(ws3, r, 10); r += 1
    for step, title, desc, fc, hc, bc in [
        ("STEP 1","アプリにアクセスする","ブラウザでアプリのURLを開き、ユーザー名・パスワードを入力してログインします。","1E40AF","DBEAFE","EFF6FF"),
        ("STEP 2","対象月を確認する","画面上部の「年・月」が自動的に当月に設定されます。\n別の月を対象にする場合は、数字を直接変更してください。","065F46","D1FAE5","F0FDF4"),
        ("STEP 3","曜日別の勤務時間を設定する","月曜〜金曜それぞれの開始時間・終了時間・休憩時間を入力します。\n設定内容はブラウザに自動保存されるため、翌月以降は変更がある場合のみ修正すれば問題ありません。","92400E","FEF3C7","FFFBEB"),
        ("STEP 4","例外日を設定する（任意）","残業・早退・休日出勤など、曜日別設定と異なる勤務時間の日がある場合は「＋ 例外日を追加」から個別に設定します。\n土日・祝日の休日出勤にも対応しています。例外日の設定は他の設定よりも最優先で適用されます。","7C3AED","EDE9FE","FAF5FF"),
        ("STEP 5","有給取得日を選択する（任意）","有給休暇を取得する日がある場合は「有給を取得する日がある」にチェックを入れ、表示されるカレンダーから該当日を選択します。\n土日・祝日も選択可能です。","BE185D","FCE7F3","FFF0F7"),
        ("STEP 6","Excelファイルを選択する","「Excelファイルを選択」エリアをクリックし、会社から配布された作業報告書（.xlsx）を選択します。\n誤ったファイルを選択した場合は、右側の「✕」ボタンでキャンセルして選び直せます。","0F766E","CCFBF1","F0FDFA"),
        ("STEP 7","入力完了・ダウンロード","「入力完了・ダウンロード」ボタンをクリックすると、自動入力済みのExcelファイルがダウンロードされます。\nダウンロードしたファイルを開いて内容を確認の上、所定の保存場所に保存してください。","1E40AF","DBEAFE","EFF6FF"),
    ]:
        _rh(ws3, r, 20)
        sc = ws3.cell(r, 2, step)
        sc.font = Font(name="Meiryo", bold=True, size=9, color="FFFFFF")
        sc.fill = _fill(fc); sc.alignment = _align("center"); sc.border = _border(fc)
        ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        tc = ws3.cell(r, 3, title)
        tc.font = Font(name="Meiryo", bold=True, size=11, color=fc)
        tc.fill = _fill(hc); tc.alignment = _align(); tc.border = _border(hc)
        ws3.cell(r, 4).fill = _fill(hc); ws3.cell(r, 4).border = _border(hc); r += 1
        lines = desc.count('\n') + 1; _rh(ws3, r, max(36, lines * 26))
        ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        dc = ws3.cell(r, 2, desc)
        dc.font = _font(size=10); dc.fill = _fill(bc)
        dc.alignment = _align(wrap=True); dc.border = _border("E5E7EB"); r += 1
        _sp(ws3, r, 5); r += 1

    # ── 自動入力ルール ──
    ws4 = wb.create_sheet("3. 自動入力ルール")
    _cw(ws4, [2, 20, 30, 30, 2]); ws4.sheet_view.showGridLines = False
    r = 1; _sp(ws4, r); r += 1
    _mg(ws4, r, 1, r, 5); _rh(ws4, r, 30)
    c = ws4.cell(r, 1, "3.  自動入力ルール")
    c.font = _font(bold=True, size=16, color="1A56DB"); c.alignment = _align("left", "center"); r += 1
    _mg(ws4, r, 1, r, 5); _rh(ws4, r, 18)
    c = ws4.cell(r, 1, "日の種類に応じて、以下のルールに従い自動で各セルへ入力されます。")
    c.font = _font(size=9, color="94A3B8", italic=True); c.alignment = _align("left", "center"); r += 1
    _sp(ws4, r, 10); r += 1
    _rh(ws4, r, 26)
    for col, label in zip([2,3,4], ["日の種類","開始・終了・休憩時間","備考欄の記入内容"]):
        hc = ws4.cell(r, col, label)
        hc.font = _font(bold=True, size=10, color="FFFFFF")
        hc.fill = _fill("1A56DB"); hc.alignment = _align("center"); hc.border = _border("1A56DB")
    r += 1
    for kind, times, note, bg, hbg in [
        ("出勤日（平日）","曜日別設定の開始・終了時間\n休憩時間：1時間00分","在宅勤務（出勤日の記入内容は変更可能）","FFFFFF","EFF6FF"),
        ("例外日（残業・早退・休日出勤）","例外日に設定した開始・終了時間\n休憩時間：1時間00分","在宅勤務（例外設定が他のすべての設定より優先）","FFFBEB","FEF3C7"),
        ("有給取得日","空欄","私用により、休暇","FFF0F7","FCE7F3"),
        ("祝日","空欄","祝日","FFF5F5","FEE2E2"),
        ("土日","空欄","空欄","F9FAFB","F3F4F6"),
    ]:
        lines = max(times.count('\n')+1, note.count('\n')+1); _rh(ws4, r, max(36, lines*24))
        kc = ws4.cell(r, 2, kind)
        kc.font = _font(bold=True, size=10); kc.fill = _fill(hbg)
        kc.alignment = _align("center"); kc.border = _border("E5E7EB")
        tc = ws4.cell(r, 3, times)
        tc.font = _font(size=10); tc.fill = _fill(bg)
        tc.alignment = _align("center", "center", wrap=True); tc.border = _border("E5E7EB")
        nc = ws4.cell(r, 4, note)
        nc.font = _font(size=10); nc.fill = _fill(bg)
        nc.alignment = _align("center", "center", wrap=True); nc.border = _border("E5E7EB"); r += 1
    _sp(ws4, r, 12); r += 1
    sec_head(ws4, r, "補足事項", "64748B"); r += 1
    for note in [
        "・祝日情報は外部API（jpholiday）から自動取得します。インターネット接続がない環境では祝日が空欄になる場合があります。",
        "・例外日の設定は、有給・祝日・土日の区別に関わらず最優先で適用されます（休日出勤の記録に使用可能）。",
        "・Excelファイルのヘッダー行を自動検索するため、列の位置が変わっても正常に動作します。",
    ]:
        _rh(ws4, r, 28); _mg(ws4, r, 2, r, 4)
        nc = ws4.cell(r, 2, note)
        nc.font = _font(size=9, color="475569"); nc.fill = _fill("F8FAFC")
        nc.alignment = _align(wrap=True); nc.border = _border("E2E8F0"); r += 1

    # ── よくある質問 ──
    ws5 = wb.create_sheet("4. よくある質問")
    _cw(ws5, [2, 6, 60, 6, 2]); ws5.sheet_view.showGridLines = False
    r = 1; _sp(ws5, r); r += 1
    _mg(ws5, r, 1, r, 5); _rh(ws5, r, 30)
    c = ws5.cell(r, 1, "4.  よくある質問（Q&A）")
    c.font = _font(bold=True, size=16, color="1A56DB"); c.alignment = _align("left", "center"); r += 1
    _sp(ws5, r, 10); r += 1
    for i, (q, a) in enumerate([
        ("設定は毎月入力し直す必要がありますか？",
         "いいえ、不要です。曜日別の勤務時間・各種ラベル設定はブラウザに自動保存されます。次回起動時には前回の設定が自動的に読み込まれるため、変更がない場合はそのままご利用いただけます。"),
        ("休日出勤が発生した場合はどのように入力しますか？",
         "「＋ 例外日を追加」ボタンから対象日を選択し（土曜・日曜・祝日も選択可）、出勤時間を入力してください。例外日の設定は曜日別設定・有給・祝日より最優先で適用されます。"),
        ("有給を土日や祝日に取得した場合はどうなりますか？",
         "有給取得日カレンダーから土日・祝日も選択可能です。選択した日は備考欄に「私用により、休暇」が記入されます。"),
        ("Excelのファイルを誤って選択した場合はどうすればよいですか？",
         "ファイル名の右側に表示される「✕」ボタンをクリックすると選択がキャンセルされます。その後、再度クリックして正しいファイルを選び直してください。"),
        ("Excelファイルの列構成が異なる場合でも使用できますか？",
         "はい、使用できます。アプリは開始時間・終了時間・休憩時間・備考のヘッダー文字列を自動検索して対象列を特定するため、列の位置が変更された場合でも正常に動作します。"),
        ("アプリはどこで動作しますか？インストールが必要ですか？",
         "ブラウザのみで動作します。インストール・追加ソフトウェアは一切不要です。アプリが起動しているサーバーにアクセスできるネットワーク環境であれば、別のPCからも利用可能です。"),
    ]):
        _rh(ws5, r, 22); _mg(ws5, r, 2, r, 4)
        qc = ws5.cell(r, 2, f"Q{i+1}　{q}")
        qc.font = _font(bold=True, size=10, color="1E40AF")
        qc.fill = _fill("DBEAFE"); qc.alignment = _align(); qc.border = _border("BFDBFE"); r += 1
        lines = a.count('。') + 1; _rh(ws5, r, max(40, lines*22))
        _mg(ws5, r, 2, r, 4)
        ac = ws5.cell(r, 2, f"A.　{a}")
        ac.font = _font(size=10); ac.fill = _fill("FFFFFF")
        ac.alignment = _align(wrap=True); ac.border = _border("E5E7EB"); r += 1
        _sp(ws5, r, 6); r += 1

    for ws in [ws1, ws2, ws3, ws4, ws5]:
        ws.oddFooter.center.text = f"日報自動入力アプリ 導入説明資料　{TODAY}　作成"
        ws.oddFooter.center.size = 8
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


@app.route('/api/report')
@auth.login_required
def download_report():
    buf = _build_report()
    return send_file(
        buf,
        as_attachment=True,
        download_name='日報アプリ_導入説明資料.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=False, port=5000)
