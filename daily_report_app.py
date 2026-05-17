import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
from datetime import date, time
import calendar as cal_module
import openpyxl
from openpyxl.styles import Font
import jpholiday

WEEKDAY_NAMES = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日']
SETTINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'settings.json')
DEFAULT_TIMES = {'start': '09:00', 'end': '17:30', 'break': '01:00'}
STANDARD_WORK_MIN = 7 * 60 + 30


def parse_time(s):
    if not s:
        return None
    try:
        h, m = s.strip().split(':')
        return time(int(h), int(m))
    except Exception:
        return None


def time_to_min(s):
    t = parse_time(s)
    if t is None:
        return None
    return t.hour * 60 + t.minute


def fmt_min(m):
    if not m:
        return ''
    return f'{m // 60}:{m % 60:02d}'


# ── カレンダーモーダル ────────────────────────────────────────────────────────

class CalendarDialog:
    """日付選択用モーダルカレンダー。"""

    def __init__(self, parent, year, month, callback):
        self.top = tk.Toplevel(parent)
        self.top.title(f'{year}年{month}月 — 日付を選択')
        self.top.resizable(False, False)
        self.top.grab_set()
        self.top.focus_set()
        self.callback = callback
        self._build(year, month)
        parent.update_idletasks()
        x = parent.winfo_rootx() + parent.winfo_width() // 2 - 165
        y = parent.winfo_rooty() + parent.winfo_height() // 2 - 150
        self.top.geometry(f'+{max(0, x)}+{max(0, y)}')

    def _build(self, year, month):
        frm = ttk.Frame(self.top, padding=14)
        frm.pack()

        ttk.Label(frm, text=f'{year}年{month}月', font=('', 11, 'bold')).grid(
            row=0, column=0, columnspan=7, pady=(0, 8))

        hdr_names = ['月', '火', '水', '木', '金', '土', '日']
        hdr_fgs   = ['#333'] * 5 + ['#3182ce', '#e53e3e']
        for c, (h, fg) in enumerate(zip(hdr_names, hdr_fgs)):
            tk.Label(frm, text=h, width=5, fg=fg, font=('', 9, 'bold'),
                     bg='#eff6ff').grid(row=1, column=c, padx=1, pady=2, sticky='ew')

        _, last = cal_module.monthrange(year, month)
        first_wd = date(year, month, 1).weekday()
        gr, gc = 2, first_wd

        for day in range(1, last + 1):
            d_obj    = date(year, month, day)
            is_sat   = gc == 5
            is_sun   = gc == 6
            hol_name = jpholiday.is_holiday_name(d_obj)
            fg = '#3182ce' if is_sat else ('#e53e3e' if (is_sun or hol_name) else '#333')
            bg = '#fff5f5' if (is_sun or hol_name) else ('#f0f8ff' if is_sat else '#ffffff')
            text = f'{day}\n{hol_name[:4]}' if hol_name else str(day)

            btn = tk.Button(frm, text=text, width=5, fg=fg, bg=bg,
                            relief='flat', cursor='hand2', font=('', 9),
                            command=lambda d=day: self._select(d))
            btn.grid(row=gr, column=gc, padx=1, pady=1, sticky='ew')
            btn.bind('<Enter>', lambda e, b=btn: b.config(bg='#dbeafe'))
            btn.bind('<Leave>', lambda e, b=btn, c=bg: b.config(bg=c))

            gc += 1
            if gc > 6:
                gc, gr = 0, gr + 1

        ttk.Button(frm, text='閉じる', command=self.top.destroy).grid(
            row=gr + 1, column=0, columnspan=7, sticky='ew', pady=(10, 0))

    def _select(self, day):
        self.callback(day)
        self.top.destroy()


# ── メインアプリ ─────────────────────────────────────────────────────────────

class DailyReportApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title('日報自動入力アプリ')
        self.root.geometry('660x880')
        self.root.minsize(580, 520)
        self._initializing = True
        self.settings = self._load_settings()
        self._build_ui()
        self._apply_settings()
        self._refresh_schedule()
        self.root.after(700, lambda: setattr(self, '_initializing', False))
        self.root.mainloop()

    # ── 設定の保存・読み込み ──────────────────────────────────────────────────

    def _load_settings(self):
        try:
            with open(SETTINGS_FILE, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

    def _save_settings(self):
        data = {
            'weekday_times': {
                str(i): {k: self.time_vars[i][k].get() for k in ('start', 'end', 'break')}
                for i in range(5)
            },
            'labels': {
                'start': self.label_start_var.get(),
                'end':   self.label_end_var.get(),
                'break': self.label_break_var.get(),
                'note':  self.label_note_var.get(),
            },
            'note_workday': self.note_workday_var.get(),
            'same_note':    self.same_note_var.get(),
        }
        try:
            with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showwarning('警告', f'設定の保存に失敗しました:\n{e}')

    def _apply_settings(self):
        wt = self.settings.get('weekday_times', {})
        for i in range(5):
            t = wt.get(str(i), DEFAULT_TIMES)
            for k in ('start', 'end', 'break'):
                self.time_vars[i][k].set(t.get(k, DEFAULT_TIMES[k]))
        labels = self.settings.get('labels', {})
        self.label_start_var.set(labels.get('start', '開始時間'))
        self.label_end_var.set(labels.get('end',    '終了時間'))
        self.label_break_var.set(labels.get('break', '休憩時間'))
        self.label_note_var.set(labels.get('note',   '備考'))
        self.note_workday_var.set(self.settings.get('note_workday', '在宅勤務'))
        self.same_note_var.set(self.settings.get('same_note', True))
        self._on_same_note_change_silent()

    # ── UI 構築 ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        canvas = tk.Canvas(self.root, borderwidth=0, highlightthickness=0)
        sb = ttk.Scrollbar(self.root, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        container = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=container, anchor='nw')
        container.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.bind_all('<MouseWheel>', lambda e: canvas.yview_scroll(-1 * (e.delta // 120), 'units'))

        pad = dict(padx=12, pady=5)

        # タイトル行：タイトル + 自動保存通知 + リセットボタン
        title_row = ttk.Frame(container)
        title_row.pack(fill='x', padx=12, pady=(14, 4))
        ttk.Label(title_row, text='日報自動入力アプリ', font=('', 14, 'bold')).pack(side='left')
        ttk.Button(title_row, text='↩ 初期値にリセット', command=self._reset_settings).pack(side='right', padx=4)
        self.autosave_notice_var = tk.StringVar(value='')
        ttk.Label(title_row, textvariable=self.autosave_notice_var,
                  foreground='#2d7a2d').pack(side='right', padx=8)

        self._build_ym(container, pad)
        self._build_weekday(container, pad)
        self._build_exceptions(container, pad)
        self._build_paid_leave(container, pad)
        self._build_schedule(container, pad)
        self._build_labels(container, pad)
        self._build_file(container, pad)
        ttk.Button(container, text='入力完了・保存', command=self._execute, width=22).pack(pady=14)

    # ── 対象月 ───────────────────────────────────────────────────────────────

    def _build_ym(self, parent, pad):
        f = ttk.LabelFrame(parent, text='対象月')
        f.pack(fill='x', **pad)
        today = date.today()
        self.year_var  = tk.IntVar(value=today.year)
        self.month_var = tk.IntVar(value=today.month)
        inner = ttk.Frame(f)
        inner.pack(padx=10, pady=6)
        ttk.Spinbox(inner, from_=2000, to=2099, textvariable=self.year_var,  width=6).pack(side='left')
        ttk.Label(inner, text=' 年 ').pack(side='left')
        ttk.Spinbox(inner, from_=1,    to=12,   textvariable=self.month_var, width=4).pack(side='left')
        ttk.Label(inner, text=' 月').pack(side='left')
        self.year_var.trace_add('write',  lambda *_: self._on_ym_change())
        self.month_var.trace_add('write', lambda *_: self._on_ym_change())

    def _on_ym_change(self):
        if self.pl_var.get():
            for w in self.pl_content.winfo_children():
                w.destroy()
            self.day_check_vars.clear()
            try:
                self._build_cal_grid(self.year_var.get(), self.month_var.get())
            except Exception:
                pass
        self._schedule_refresh()

    # ── 曜日別勤務時間 ────────────────────────────────────────────────────────

    def _build_weekday(self, parent, pad):
        f = ttk.LabelFrame(parent, text='曜日別勤務時間  （形式: HH:MM）')
        f.pack(fill='x', **pad)
        for col, txt in enumerate(['', '開始', '終了', '休憩']):
            ttk.Label(f, text=txt, width=8 if col else 7, anchor='center').grid(
                row=0, column=col, padx=3, pady=2)
        self.time_vars = []
        for i, name in enumerate(WEEKDAY_NAMES):
            ttk.Label(f, text=name, width=6, anchor='w').grid(
                row=i + 1, column=0, padx=10, pady=4, sticky='w')
            row_vars = {}
            for col, key in enumerate(('start', 'end', 'break'), 1):
                var = tk.StringVar()
                var.trace_add('write', lambda *_: self._on_input_change())
                ttk.Entry(f, textvariable=var, width=8, justify='center').grid(
                    row=i + 1, column=col, padx=3, pady=4)
                row_vars[key] = var
            self.time_vars.append(row_vars)

    # ── 例外日（カレンダーで日付選択） ───────────────────────────────────────

    def _build_exceptions(self, parent, pad):
        outer = ttk.LabelFrame(parent, text='例外日（残業・早退・休日出勤など）')
        outer.pack(fill='x', **pad)

        hdr = ttk.Frame(outer)
        hdr.pack(fill='x', padx=8, pady=(4, 0))
        for col, (txt, w) in enumerate([('日付', 11), ('開始', 7), ('終了', 7), ('休憩', 7), ('備考', 14)]):
            ttk.Label(hdr, text=txt, width=w, anchor='center').grid(row=0, column=col, padx=2)

        self.ex_rows_frame = ttk.Frame(outer)
        self.ex_rows_frame.pack(fill='x', padx=8)

        ttk.Button(outer, text='＋ 例外日を追加', command=self._add_ex_row).pack(pady=5)
        self.exception_rows = []

    def _add_ex_row(self):
        rf = ttk.Frame(self.ex_rows_frame)
        rf.pack(fill='x', pady=2)

        day_v   = tk.StringVar(value='')
        start_v = tk.StringVar()
        end_v   = tk.StringVar()
        break_v = tk.StringVar(value='01:00')
        note_v  = tk.StringVar(value=self.note_workday_var.get() if hasattr(self, 'note_workday_var') else '在宅勤務')

        for v in (start_v, end_v, break_v, note_v):
            v.trace_add('write', lambda *_: self._schedule_refresh())

        date_btn = ttk.Button(rf, text='日付を選択', width=11)
        date_btn.grid(row=0, column=0, padx=2)

        row_data = {
            'frame': rf, 'day': day_v, 'date_btn': date_btn,
            'start': start_v, 'end': end_v, 'break': break_v, 'note': note_v,
        }

        def open_cal(r=row_data):
            def on_select(d, r=r):
                r['day'].set(str(d))
                m = self.month_var.get()
                r['date_btn'].config(text=f'{m}月{d}日')
                self._schedule_refresh()
            CalendarDialog(self.root, self.year_var.get(), self.month_var.get(), on_select)

        date_btn.config(command=open_cal)

        ttk.Entry(rf, textvariable=start_v, width=7, justify='center').grid(row=0, column=1, padx=2)
        ttk.Entry(rf, textvariable=end_v,   width=7, justify='center').grid(row=0, column=2, padx=2)
        ttk.Entry(rf, textvariable=break_v, width=7, justify='center').grid(row=0, column=3, padx=2)
        ttk.Entry(rf, textvariable=note_v,  width=14).grid(row=0, column=4, padx=2)

        def remove(r=row_data):
            r['frame'].destroy()
            self.exception_rows.remove(r)
            self._schedule_refresh()

        ttk.Button(rf, text='✕', width=2, command=remove).grid(row=0, column=5, padx=4)
        self.exception_rows.append(row_data)

    # ── 有給取得日 ───────────────────────────────────────────────────────────

    def _build_paid_leave(self, parent, pad):
        outer = ttk.LabelFrame(parent, text='有給取得日')
        outer.pack(fill='x', **pad)
        self.pl_var = tk.BooleanVar()
        ttk.Checkbutton(outer, text='有給を取得する日がある', variable=self.pl_var,
                        command=self._toggle_pl).pack(anchor='w', padx=8, pady=4)
        self.pl_content = ttk.Frame(outer)
        self.day_check_vars = {}

    def _toggle_pl(self):
        for w in self.pl_content.winfo_children():
            w.destroy()
        self.day_check_vars.clear()
        if self.pl_var.get():
            self.pl_content.pack(padx=10, pady=(0, 8))
            self._build_cal_grid(self.year_var.get(), self.month_var.get())
        else:
            self.pl_content.pack_forget()
        self._schedule_refresh()

    def _build_cal_grid(self, year, month):
        _, last = cal_module.monthrange(year, month)
        hdr_names = ['月', '火', '水', '木', '金', '土', '日']
        hdr_fgs   = ['black'] * 5 + ['royalblue', 'red']
        for c, (h, fg) in enumerate(zip(hdr_names, hdr_fgs)):
            tk.Label(self.pl_content, text=h, width=4, fg=fg, font=('', 9, 'bold')).grid(
                row=0, column=c, padx=1)
        first_wd = date(year, month, 1).weekday()
        gr, gc = 1, first_wd
        for day in range(1, last + 1):
            var = tk.BooleanVar()
            var.trace_add('write', lambda *_: self._schedule_refresh())
            self.day_check_vars[day] = var
            fg = 'royalblue' if gc == 5 else ('red' if gc == 6 else 'black')
            tk.Checkbutton(self.pl_content, text=str(day), variable=var, fg=fg, width=3).grid(
                row=gr, column=gc, padx=1, pady=1)
            gc += 1
            if gc > 6:
                gc, gr = 0, gr + 1

    # ── 月間スケジュール・合計稼働時間 ───────────────────────────────────────

    def _build_schedule(self, parent, pad):
        f = ttk.LabelFrame(parent, text='合計稼働時間 / 月間スケジュール')
        f.pack(fill='x', **pad)

        top = ttk.Frame(f)
        top.pack(fill='x', padx=8, pady=(6, 4))
        ttk.Label(top, text='合計稼働時間:').pack(side='left')
        self.total_hours_var = tk.StringVar(value='--')
        ttk.Label(top, textvariable=self.total_hours_var,
                  font=('', 13, 'bold'), foreground='#1a56db').pack(side='left', padx=8)
        ttk.Button(top, text='↺ 更新', command=self._refresh_schedule).pack(side='right')

        cols = ('日', '曜', '開始', '終了', '休憩', '就業時間', '備考')
        tree_frame = ttk.Frame(f)
        tree_frame.pack(fill='x', padx=8, pady=(0, 8))

        self.schedule_tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=12)
        widths = {'日': 28, '曜': 28, '開始': 58, '終了': 58, '休憩': 58, '就業時間': 72, '備考': 130}
        for col in cols:
            self.schedule_tree.heading(col, text=col)
            self.schedule_tree.column(col, width=widths[col],
                                      anchor='w' if col == '備考' else 'center',
                                      stretch=col == '備考')

        self.schedule_tree.tag_configure('sun',     foreground='#e53e3e', background='#fff5f5')
        self.schedule_tree.tag_configure('sat',     foreground='#3182ce', background='#f0f8ff')
        self.schedule_tree.tag_configure('holiday', foreground='#e53e3e', background='#fff5f5')
        self.schedule_tree.tag_configure('paid',    background='#faf5ff')
        self.schedule_tree.tag_configure('exc',     background='#fffbeb')
        self.schedule_tree.tag_configure('work',    background='#ffffff')

        ysb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.schedule_tree.yview)
        self.schedule_tree.configure(yscrollcommand=ysb.set)
        self.schedule_tree.pack(side='left', fill='both', expand=True)
        ysb.pack(side='right', fill='y')

    def _schedule_refresh(self):
        if hasattr(self, '_refresh_id'):
            self.root.after_cancel(self._refresh_id)
        self._refresh_id = self.root.after(300, self._refresh_schedule)

    def _refresh_schedule(self):
        if not hasattr(self, 'schedule_tree'):
            return
        try:
            year  = self.year_var.get()
            month = self.month_var.get()
        except Exception:
            return
        if not (2000 <= year <= 2099 and 1 <= month <= 12):
            return

        _, last = cal_module.monthrange(year, month)
        DOW_NAMES = ['月', '火', '水', '木', '金', '土', '日']

        wt = {}
        for i in range(5):
            s = time_to_min(self.time_vars[i]['start'].get())
            e = time_to_min(self.time_vars[i]['end'].get())
            b = time_to_min(self.time_vars[i]['break'].get()) or 0
            if s is not None and e is not None and e > s:
                wt[i] = {
                    'start': self.time_vars[i]['start'].get(),
                    'end':   self.time_vars[i]['end'].get(),
                    'break': self.time_vars[i]['break'].get(),
                    'work':  e - s - b,
                }

        paid = {d for d, v in self.day_check_vars.items() if v.get()} if self.pl_var.get() else set()

        tex = {}
        for row in self.exception_rows:
            ds = row['day'].get()
            if not (ds and ds.isdigit()):
                continue
            d = int(ds)
            s = time_to_min(row['start'].get())
            e = time_to_min(row['end'].get())
            b = time_to_min(row['break'].get()) or 0
            if s is not None and e is not None and e > s:
                tex[d] = {
                    'start': row['start'].get(), 'end': row['end'].get(),
                    'break': row['break'].get(), 'note': row['note'].get(),
                    'work':  e - s - b,
                }

        same_note    = self.same_note_var.get()
        note_workday = self.note_workday_var.get().strip() or '在宅勤務'

        # 日付別備考（same_note=False のとき使用）
        nex = {}
        if not same_note and hasattr(self, 'note_exception_rows'):
            for row in self.note_exception_rows:
                ds = row['day'].get()
                if ds and ds.isdigit():
                    n = row['note'].get().strip()
                    if n:
                        nex[int(ds)] = n

        hols = {}
        for d in range(1, last + 1):
            name = jpholiday.is_holiday_name(date(year, month, d))
            if name:
                hols[d] = name

        for item in self.schedule_tree.get_children():
            self.schedule_tree.delete(item)

        total_work = 0

        for day in range(1, last + 1):
            wd       = date(year, month, day).weekday()
            dow_name = DOW_NAMES[wd]
            start = end = brk = work_str = note = ''
            tag = 'work'

            if day in tex:
                t = tex[day]
                start, end, brk = t['start'], t['end'], t['break']
                work_str = fmt_min(t['work'])
                total_work += t['work']
                if same_note:
                    note = note_workday
                else:
                    note = t['note'] or nex.get(day, note_workday)
                tag = 'exc'
            elif wd == 6:
                tag = 'sun'
            elif wd == 5:
                tag = 'sat'
            elif day in hols:
                note = '祝日'
                tag = 'holiday'
            elif day in paid:
                note = '私用により、休暇'
                tag = 'paid'
            elif wd in wt:
                t = wt[wd]
                start, end, brk = t['start'], t['end'], t['break']
                work_str = fmt_min(t['work'])
                total_work += t['work']
                note = note_workday if same_note else nex.get(day, note_workday)
                tag = 'work'

            self.schedule_tree.insert('', 'end',
                values=(day, dow_name, start, end, brk, work_str, note), tags=(tag,))

        self.total_hours_var.set(fmt_min(total_work) or '--')

    # ── Excel列ラベル / 備考設定 ──────────────────────────────────────────────

    def _build_labels(self, parent, pad):
        f = ttk.LabelFrame(parent, text='Excel列ラベル / 備考設定')
        f.pack(fill='x', **pad)

        self.label_start_var  = tk.StringVar(value='開始時間')
        self.label_end_var    = tk.StringVar(value='終了時間')
        self.label_break_var  = tk.StringVar(value='休憩時間')
        self.label_note_var   = tk.StringVar(value='備考')
        self.note_workday_var = tk.StringVar(value='在宅勤務')
        self.same_note_var    = tk.BooleanVar(value=True)

        inner = ttk.Frame(f)
        inner.pack(fill='x', padx=10, pady=6)

        lbl_row = ttk.Frame(inner)
        lbl_row.pack(fill='x', pady=2)
        for txt, var, w in [
            ('開始時間列', self.label_start_var,  9),
            ('終了時間列', self.label_end_var,    9),
            ('休憩時間列', self.label_break_var,  9),
            ('備考列',     self.label_note_var,   7),
        ]:
            col = ttk.Frame(lbl_row)
            col.pack(side='left', padx=6)
            ttk.Label(col, text=txt, font=('', 8)).pack()
            ttk.Entry(col, textvariable=var, width=w).pack()

        note_row = ttk.Frame(inner)
        note_row.pack(fill='x', pady=(8, 2))
        ttk.Label(note_row, text='出勤日の備考:').pack(side='left')
        ttk.Entry(note_row, textvariable=self.note_workday_var, width=18).pack(side='left', padx=6)
        ttk.Checkbutton(note_row, text='出勤日はすべて同じ',
                        variable=self.same_note_var,
                        command=self._on_same_note_change).pack(side='left', padx=4)

        # 日付別備考フレーム（same_note=False のとき表示）
        self.note_ex_frame = ttk.Frame(inner)

        nex_hdr = ttk.Frame(self.note_ex_frame)
        nex_hdr.pack(fill='x', pady=(6, 2))
        ttk.Label(nex_hdr, text='日付別備考（出勤日）:', font=('', 8)).pack(side='left')
        ttk.Button(nex_hdr, text='＋ 追加', command=self._add_note_ex_row).pack(side='left', padx=6)

        self.note_ex_rows_frame = ttk.Frame(self.note_ex_frame)
        self.note_ex_rows_frame.pack(fill='x')
        self.note_exception_rows = []

        self.note_workday_var.trace_add('write', lambda *_: self._on_input_change())
        for var in (self.label_start_var, self.label_end_var, self.label_break_var, self.label_note_var):
            var.trace_add('write', lambda *_: self._on_input_change())

    def _on_same_note_change(self):
        self._on_same_note_change_silent()
        self._on_input_change()

    def _on_same_note_change_silent(self):
        if self.same_note_var.get():
            self.note_ex_frame.pack_forget()
        else:
            self.note_ex_frame.pack(fill='x', pady=(4, 0))
            if hasattr(self, 'note_exception_rows') and not self.note_exception_rows:
                self._add_note_ex_row()

    def _add_note_ex_row(self):
        rf = ttk.Frame(self.note_ex_rows_frame)
        rf.pack(fill='x', pady=2)

        day_v  = tk.StringVar(value='')
        note_v = tk.StringVar()
        note_v.trace_add('write', lambda *_: self._schedule_refresh())

        date_btn = ttk.Button(rf, text='日付を選択', width=11)
        date_btn.grid(row=0, column=0, padx=2)

        row_data = {'frame': rf, 'day': day_v, 'date_btn': date_btn, 'note': note_v}

        def open_cal(r=row_data):
            def on_select(d, r=r):
                r['day'].set(str(d))
                m = self.month_var.get()
                r['date_btn'].config(text=f'{m}月{d}日')
                self._schedule_refresh()
            CalendarDialog(self.root, self.year_var.get(), self.month_var.get(), on_select)

        date_btn.config(command=open_cal)
        ttk.Entry(rf, textvariable=note_v, width=24).grid(row=0, column=1, padx=4, sticky='ew')

        def remove(r=row_data):
            r['frame'].destroy()
            self.note_exception_rows.remove(r)
            self._schedule_refresh()

        ttk.Button(rf, text='✕', width=2, command=remove).grid(row=0, column=2, padx=2)
        self.note_exception_rows.append(row_data)

    # ── Excelファイル選択 ─────────────────────────────────────────────────────

    def _build_file(self, parent, pad):
        f = ttk.LabelFrame(parent, text='Excelファイル')
        f.pack(fill='x', **pad)
        inner = ttk.Frame(f)
        inner.pack(fill='x', padx=8, pady=6)
        inner.columnconfigure(0, weight=1)

        self.file_var = tk.StringVar()
        ttk.Entry(inner, textvariable=self.file_var, width=44).grid(
            row=0, column=0, sticky='ew', padx=(0, 4))
        ttk.Button(inner, text='参照...', command=self._browse).grid(row=0, column=1, padx=2)
        self.clear_file_btn = ttk.Button(inner, text='✕', width=2, command=self._clear_file)

        self.file_var.trace_add('write', lambda *_: self._update_clear_btn())

    def _browse(self):
        p = filedialog.askopenfilename(
            title='Excelファイルを選択してください',
            filetypes=[('Excelファイル', '*.xlsx'), ('すべてのファイル', '*.*')],
        )
        if p:
            self.file_var.set(p)

    def _update_clear_btn(self):
        if self.file_var.get():
            self.clear_file_btn.grid(row=0, column=2, padx=2)
        else:
            self.clear_file_btn.grid_remove()

    def _clear_file(self):
        self.file_var.set('')

    # ── 自動保存・通知・リセット ──────────────────────────────────────────────

    def _reset_settings(self):
        for i in range(5):
            for k, v in DEFAULT_TIMES.items():
                self.time_vars[i][k].set(v)
        self.label_start_var.set('開始時間')
        self.label_end_var.set('終了時間')
        self.label_break_var.set('休憩時間')
        self.label_note_var.set('備考')
        self.note_workday_var.set('在宅勤務')
        self.same_note_var.set(True)
        self._on_same_note_change_silent()
        for row in self.exception_rows[:]:
            row['frame'].destroy()
        self.exception_rows.clear()
        for row in self.note_exception_rows[:]:
            row['frame'].destroy()
        self.note_exception_rows.clear()
        self.pl_var.set(False)
        self._toggle_pl()
        self._show_notice('↩ リセットしました')
        self._refresh_schedule()

    def _show_notice(self, text):
        self.autosave_notice_var.set(text)
        if hasattr(self, '_notice_id'):
            self.root.after_cancel(self._notice_id)
        self._notice_id = self.root.after(2000, lambda: self.autosave_notice_var.set(''))

    def _schedule_auto_save(self):
        if self._initializing:
            return
        if hasattr(self, '_autosave_id'):
            self.root.after_cancel(self._autosave_id)
        self._autosave_id = self.root.after(500, self._do_auto_save)

    def _do_auto_save(self):
        self._save_settings()
        self._show_notice('✓ 自動保存しました')

    def _on_input_change(self):
        self._schedule_refresh()
        self._schedule_auto_save()

    # ── 書き込み処理 ──────────────────────────────────────────────────────────

    def _execute(self):
        path = self.file_var.get().strip()
        if not path:
            messagebox.showerror('エラー', 'Excelファイルを選択してください。')
            return
        if not os.path.exists(path):
            messagebox.showerror('エラー', 'ファイルが見つかりません。\nパスを確認してください。')
            return

        year, month = self.year_var.get(), self.month_var.get()

        wt = {}
        for i in range(5):
            s = parse_time(self.time_vars[i]['start'].get())
            e = parse_time(self.time_vars[i]['end'].get())
            b = parse_time(self.time_vars[i]['break'].get()) or time(1, 0)
            if s and e:
                wt[i] = {'start': s, 'end': e, 'break': b}
        if not wt:
            messagebox.showerror('エラー', '勤務時間をHH:MM形式で入力してください。\n例: 09:00')
            return

        paid = {d for d, v in self.day_check_vars.items() if v.get()} if self.pl_var.get() else set()

        tex = {}
        for row in self.exception_rows:
            ds = row['day'].get()
            if not (ds and ds.isdigit()):
                continue
            d = int(ds)
            s = parse_time(row['start'].get())
            e = parse_time(row['end'].get())
            b = parse_time(row['break'].get()) or time(1, 0)
            if s and e:
                tex[d] = {'start': s, 'end': e, 'break': b, 'note': row['note'].get().strip()}

        same_note    = self.same_note_var.get()
        note_workday = self.note_workday_var.get().strip() or '在宅勤務'

        nex = {}
        if not same_note:
            for row in self.note_exception_rows:
                ds = row['day'].get()
                if ds and ds.isdigit():
                    n = row['note'].get().strip()
                    if n:
                        nex[int(ds)] = n

        _, last = cal_module.monthrange(year, month)
        hols = {d for d in range(1, last + 1) if jpholiday.is_holiday(date(year, month, d))}

        label_start  = self.label_start_var.get().strip()  or '開始時間'
        label_end    = self.label_end_var.get().strip()    or '終了時間'
        label_break  = self.label_break_var.get().strip()  or '休憩時間'
        label_note   = self.label_note_var.get().strip()   or '備考'

        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            messagebox.showerror('エラー', f'ファイルを開けませんでした。\n{e}')
            return
        ws = wb.active

        hr = self._find_hr(ws, [label_start, label_end, label_break, label_note]) or 19
        dr = hr + 1
        cs = self._find_col(ws, hr, label_start) or 'F'
        ce = self._find_col(ws, hr, label_end)   or 'I'
        cb = self._find_col(ws, hr, label_break) or 'L'
        cn = self._find_col(ws, hr, label_note)  or 'S'
        blk = Font(color='000000')

        for day in range(1, last + 1):
            r  = dr + day - 1
            wd = date(year, month, day).weekday()

            for c in (cs, ce, cb, cn):
                ws[f'{c}{r}'].value = None

            if day in tex:
                t = tex[day]
                for c, v in ((cs, t['start']), (ce, t['end']), (cb, t['break'])):
                    ws[f'{c}{r}'].value         = v
                    ws[f'{c}{r}'].number_format = 'h:mm'
                    ws[f'{c}{r}'].font          = blk
                if same_note:
                    note = note_workday
                else:
                    note = t['note'] or nex.get(day, note_workday)
                ws[f'{cn}{r}'].value = note
                ws[f'{cn}{r}'].font  = blk
            elif day in paid:
                ws[f'{cn}{r}'].value = '私用により、休暇'
                ws[f'{cn}{r}'].font  = blk
            elif wd >= 5:
                pass
            elif day in hols:
                ws[f'{cn}{r}'].value = '祝日'
                ws[f'{cn}{r}'].font  = blk
            elif wd in wt:
                t = wt[wd]
                for c, v in ((cs, t['start']), (ce, t['end']), (cb, t['break'])):
                    ws[f'{c}{r}'].value         = v
                    ws[f'{c}{r}'].number_format = 'h:mm'
                    ws[f'{c}{r}'].font          = blk
                note = note_workday if same_note else nex.get(day, note_workday)
                ws[f'{cn}{r}'].value = note
                ws[f'{cn}{r}'].font  = blk

        try:
            ws.sheet_view.selection[0].activeCell = 'A1'
            ws.sheet_view.selection[0].sqref      = 'A1'
        except Exception:
            pass

        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror('エラー', f'保存に失敗しました。\nExcelが開かれていないか確認してください。\n{e}')
            return

        self._save_settings()
        messagebox.showinfo('完了', f'書き込みが完了しました。\n\nファイル: {os.path.basename(path)}')

    # ── Excel ヘルパー ────────────────────────────────────────────────────────

    def _find_col(self, ws, hr, label):
        label = label.strip()
        for cell in ws[hr]:
            if cell.value:
                v = str(cell.value).strip()
                if v == label or label in v or v in label:
                    return cell.column_letter
        return None

    def _find_hr(self, ws, labels):
        clean = [l.strip() for l in labels if l.strip()]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            vals = [str(c.value).strip() for c in row if c.value]
            if sum(1 for lbl in clean if any(lbl in v or v in lbl for v in vals)) >= 2:
                return row[0].row
        return None


if __name__ == '__main__':
    DailyReportApp()
