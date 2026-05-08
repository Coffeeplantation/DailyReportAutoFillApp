import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "使い方ガイド"

# ── 列幅 ──
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 52
ws.column_dimensions['D'].width = 28

# ── 色定義 ──
C_NAVY   = "1A56DB"
C_LIGHT  = "EFF6FF"
C_ACCENT = "DBEAFE"
C_GREEN  = "059669"
C_GREEN2 = "D1FAE5"
C_ORANGE = "EA580C"
C_ORG2   = "FFF7ED"
C_GRAY   = "F3F4F6"
C_WHITE  = "FFFFFF"
C_BORDER = "CBD5E1"

def side(color=C_BORDER, style="thin"):
    return Side(border_style=style, color=color)

def border(left=True, right=True, top=True, bottom=True, color=C_BORDER, style="thin"):
    return Border(
        left=side(color, style) if left else Side(),
        right=side(color, style) if right else Side(),
        top=side(color, style) if top else Side(),
        bottom=side(color, style) if bottom else Side(),
    )

def fill(color):
    return PatternFill("solid", fgColor=color)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Meiryo", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def write(row, col, value, fnt=None, fll=None, aln=None, brd=None):
    c = ws.cell(row=row, column=col, value=value)
    if fnt: c.font = fnt
    if fll: c.fill = fll
    if aln: c.alignment = aln
    if brd: c.border = brd
    return c

def merge(r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def row_height(row, h):
    ws.row_dimensions[row].height = h


# ════════════════════════════════════════════════════
# タイトルブロック（行1〜4）
# ════════════════════════════════════════════════════
for r in range(1, 5):
    row_height(r, 6 if r in (1, 4) else 24)

# 背景
for r in range(1, 5):
    for c in range(1, 5):
        ws.cell(r, c).fill = fill(C_NAVY)

merge(2, 1, 3, 4)
c = ws.cell(2, 1, "📋  日報自動入力アプリ  使い方ガイド")
c.font = Font(name="Meiryo", bold=True, size=18, color=C_WHITE)
c.alignment = align("center", "center")

# ════════════════════════════════════════════════════
# セクション：アプリ概要（行6〜14）
# ════════════════════════════════════════════════════
row_height(5, 8)

# セクション見出し
merge(6, 1, 6, 4)
row_height(6, 28)
c = ws.cell(6, 1, "■  アプリ概要")
c.font = font(bold=True, size=13, color=C_WHITE)
c.fill = fill(C_NAVY)
c.alignment = align("left", "center")
c.border = border()

overview = [
    ("目的",     "毎月会社から配布されるExcel作業報告書に、勤務時間・備考を自動で書き込み・保存するWebアプリです。"),
    ("メリット", "月初に最小限の入力をするだけで、1ヶ月分のExcelが自動で完成します。手入力の手間をゼロにします。"),
    ("対象",     "Excel形式（.xlsx）の作業報告書を毎月提出している社員が対象です。"),
    ("動作環境", "ブラウザ（Chrome / Edge / Safari など）で動作します。インストール不要です。"),
]

for i, (label, text) in enumerate(overview):
    r = 7 + i
    row_height(r, 36)
    c1 = ws.cell(r, 2, label)
    c1.font = font(bold=True, size=10, color=C_NAVY)
    c1.fill = fill(C_ACCENT)
    c1.alignment = align("center", "center")
    c1.border = border()
    merge(r, 3, r, 4)
    c2 = ws.cell(r, 3, text)
    c2.font = font(size=10)
    c2.fill = fill(C_WHITE)
    c2.alignment = align("left", "center", wrap=True)
    c2.border = border()

# ════════════════════════════════════════════════════
# セクション：使い方手順（行16〜）
# ════════════════════════════════════════════════════
row_height(12, 8)

# セクション見出し
merge(13, 1, 13, 4)
row_height(13, 28)
c = ws.cell(13, 1, "■  使い方手順")
c.font = font(bold=True, size=13, color=C_WHITE)
c.fill = fill(C_GREEN)
c.alignment = align("left", "center")
c.border = border()

steps = [
    ("STEP 1", "アプリを開く",
     "ブラウザでアプリのURLにアクセスします。ユーザー名・パスワードを入力してログインしてください。"),
    ("STEP 2", "対象月を確認する",
     "画面上部に「年・月」が表示されます。自動で当月が設定されます。\n異なる月を入力したい場合は数字を直接変更してください。"),
    ("STEP 3", "曜日別の勤務時間を入力する",
     "「曜日別の勤務時間」カードで、月〜金それぞれの開始・終了・休憩時間を入力します。\n入力した設定はブラウザに自動保存され、次回起動時に引き継がれます。"),
    ("STEP 4", "例外日を設定する（任意）",
     "「＋ 例外日を追加」ボタンで、通常と異なる時間にしたい日（残業・早退・休日出勤など）を追加します。\n追加した日の時間はすべての設定より優先されます（土日・祝日の休日出勤にも対応）。"),
    ("STEP 5", "有給取得日を選択する（任意）",
     "「有給を取得する日がある」にチェックを入れるとカレンダーが表示されます。\n有給を取得する日をクリックして選択してください（土日・祝日も選択可能）。"),
    ("STEP 6", "Excelファイルを選択する",
     "「Excelファイルを選択」カードをクリックし、会社から配布された作業報告書（.xlsx）を選びます。\n選択後、✕ボタンで選択をやり直せます。"),
    ("STEP 7", "入力完了・ダウンロード",
     "「入力完了・ダウンロード」ボタンを押すと、自動入力済みのExcelファイルがダウンロードされます。\nダウンロードしたファイルを開いて内容を確認し、所定の場所に保存してください。"),
]

step_colors = [
    ("1E40AF", "DBEAFE"),
    ("065F46", "D1FAE5"),
    ("92400E", "FEF3C7"),
    ("7C3AED", "EDE9FE"),
    ("BE185D", "FCE7F3"),
    ("1E40AF", "DBEAFE"),
    ("065F46", "D1FAE5"),
]

cur_row = 14
for i, (step, title, desc) in enumerate(steps):
    fc, bc = step_colors[i]
    row_height(cur_row, 14)
    # STEP バッジ
    c1 = ws.cell(cur_row, 2, step)
    c1.font = Font(name="Meiryo", bold=True, size=9, color=C_WHITE)
    c1.fill = fill(fc)
    c1.alignment = align("center", "center")
    c1.border = border()
    # タイトル
    c2 = ws.cell(cur_row, 3, title)
    c2.font = Font(name="Meiryo", bold=True, size=11, color=fc)
    c2.fill = fill(bc)
    c2.alignment = align("left", "center")
    c2.border = border()
    ws.cell(cur_row, 4).fill = fill(bc)
    ws.cell(cur_row, 4).border = border()
    ws.merge_cells(start_row=cur_row, start_column=3, end_row=cur_row, end_column=4)

    cur_row += 1
    # 説明文
    line_count = desc.count('\n') + 1
    h = max(32, line_count * 28)
    row_height(cur_row, h)
    ws.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row, end_column=4)
    c3 = ws.cell(cur_row, 2, desc)
    c3.font = font(size=10)
    c3.fill = fill(C_WHITE)
    c3.alignment = align("left", "center", wrap=True)
    c3.border = border()

    cur_row += 1
    row_height(cur_row, 5)
    cur_row += 1

# ════════════════════════════════════════════════════
# セクション：自動入力ルール表
# ════════════════════════════════════════════════════
row_height(cur_row, 8)
cur_row += 1

merge(cur_row, 1, cur_row, 4)
row_height(cur_row, 28)
c = ws.cell(cur_row, 1, "■  自動入力ルール")
c.font = font(bold=True, size=13, color=C_WHITE)
c.fill = fill(C_ORANGE)
c.alignment = align("left", "center")
c.border = border()
cur_row += 1

# テーブルヘッダー
headers = ["日の種類", "開始・終了・休憩", "備考欄"]
header_row = cur_row
row_height(cur_row, 22)
ws.cell(cur_row, 2).fill = fill(C_ORANGE)
ws.cell(cur_row, 2).border = border()
for j, h in enumerate(headers):
    c = ws.cell(cur_row, 2 + j, h)
    c.font = font(bold=True, size=10, color=C_WHITE)
    c.fill = fill(C_ORANGE)
    c.alignment = align("center", "center")
    c.border = border()

merge(cur_row, 3, cur_row, 3)
cur_row += 1

rules = [
    ("出勤日（平日）",   "曜日別設定の時間・休憩 1:00",   "在宅勤務（変更可）",  C_WHITE),
    ("例外日（休日出勤）","例外設定した時間・休憩 1:00",   "在宅勤務（変更可）",  "FFF9C4"),
    ("有給取得日",        "空欄",                           "私用により、休暇",    "FCE7F3"),
    ("祝日",              "空欄",                           "祝日",                "FFF5F5"),
    ("土日",              "空欄",                           "空欄",                C_GRAY),
]

for kind, times, note, bg in rules:
    row_height(cur_row, 28)
    for col, val in zip([2, 3, 4], [kind, times, note]):
        c = ws.cell(cur_row, col, val)
        c.font = font(size=10)
        c.fill = fill(bg)
        c.alignment = align("center", "center")
        c.border = border()
    cur_row += 1

# ════════════════════════════════════════════════════
# セクション：よくある疑問
# ════════════════════════════════════════════════════
row_height(cur_row, 8)
cur_row += 1

merge(cur_row, 1, cur_row, 4)
row_height(cur_row, 28)
c = ws.cell(cur_row, 1, "■  よくある疑問")
c.font = font(bold=True, size=13, color=C_WHITE)
c.fill = fill("6D28D9")
c.alignment = align("left", "center")
c.border = border()
cur_row += 1

faqs = [
    ("設定は毎月入力し直す？",
     "いいえ。曜日別の勤務時間・各種ラベルはブラウザに自動保存されます。次回起動時には前回の設定が自動で反映されます。"),
    ("休日出勤はどう入力する？",
     "「＋ 例外日を追加」から該当日を選択（土日・祝日も選択可）し、出勤時間を入力します。例外日設定は最優先で反映されます。"),
    ("有給を祝日や土日に取得したい場合は？",
     "有給カレンダーで土日・祝日もクリック選択できます。選択した日は「私用により、休暇」が記入されます。"),
    ("ファイルを間違えて選択した場合は？",
     "ファイル名の右に表示される「✕」ボタンを押すと選択をキャンセルできます。再度クリックして正しいファイルを選び直してください。"),
    ("書き込まれる列や行がずれている場合は？",
     "アプリが「開始時間・終了時間・休憩時間・備考」のヘッダーを自動検索して列を特定します。Excelのヘッダー文字列と設定が一致しているか確認してください。"),
]

for q, a in faqs:
    row_height(cur_row, 18)
    merge(cur_row, 2, cur_row, 4)
    c = ws.cell(cur_row, 2, f"Q.  {q}")
    c.font = Font(name="Meiryo", bold=True, size=10, color="6D28D9")
    c.fill = fill("EDE9FE")
    c.alignment = align("left", "center")
    c.border = border()
    cur_row += 1

    row_height(cur_row, max(32, a.count('\n') * 28 + 28))
    merge(cur_row, 2, cur_row, 4)
    c = ws.cell(cur_row, 2, f"A.  {a}")
    c.font = font(size=10)
    c.fill = fill(C_WHITE)
    c.alignment = align("left", "center", wrap=True)
    c.border = border()
    cur_row += 1
    row_height(cur_row, 4)
    cur_row += 1

# ════════════════════════════════════════════════════
# フッター
# ════════════════════════════════════════════════════
row_height(cur_row, 8)
cur_row += 1
merge(cur_row, 1, cur_row, 4)
row_height(cur_row, 22)
c = ws.cell(cur_row, 1, "日報自動入力アプリ  操作マニュアル  ／  本ファイルはアプリから自動生成されました")
c.font = font(size=9, color="94A3B8", italic=True)
c.fill = fill(C_NAVY)
c.alignment = align("center", "center")

# 左端A列の背景をネイビーで統一（装飾）
for r in range(1, cur_row + 1):
    ws.cell(r, 1).fill = fill(C_NAVY)

wb.save("/workspaces/DailyReportAutoFillApp/日報アプリ_使い方ガイド.xlsx")
print("完了")
