"""上長提出用：日報自動入力アプリ 導入説明資料 生成スクリプト"""
import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = date.today().strftime("%Y年%m月%d日")

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════
# ヘルパー
# ════════════════════════════════════════════════════
def _side(c="CBD5E1"): return Side(border_style="thin", color=c)
def _border(color="CBD5E1"):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)
def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, size=11, color="1a1a1a", italic=False, name="Meiryo"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def rh(ws, row, h): ws.row_dimensions[row].height = h
def mg(ws, r1, c1, r2, c2): ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def col_widths(ws, widths):
    for col, w in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = w

def section_heading(ws, row, text, bg="1A56DB", colspan=6):
    mg(ws, row, 1, row, colspan)
    rh(ws, row, 26)
    c = ws.cell(row, 1, text)
    c.font = _font(bold=True, size=12, color="FFFFFF")
    c.fill = _fill(bg)
    c.alignment = _align("left", "center")
    c.border = _border(bg)

def body_cell(ws, row, col, value, bg="FFFFFF", bold=False, size=10,
              color="1a1a1a", h="left", wrap=True, colspan=None, border_color="CBD5E1"):
    if colspan:
        mg(ws, row, col, row, col + colspan - 1)
    c = ws.cell(row, col, value)
    c.font = _font(bold=bold, size=size, color=color)
    c.fill = _fill(bg)
    c.alignment = _align(h, "center", wrap)
    c.border = _border(border_color)
    return c

def spacer(ws, row, h=6):
    rh(ws, row, h)


# ════════════════════════════════════════════════════
# Sheet 1：表紙
# ════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "表紙"
col_widths(ws1, [4, 20, 44, 20, 4])
ws1.sheet_view.showGridLines = False

for r in range(1, 32):
    rh(ws1, r, 18)
    for c in range(1, 6):
        ws1.cell(r, c).fill = _fill("F0F4FF")

# タイトルブロック
for r in range(5, 17):
    for c in range(2, 5):
        ws1.cell(r, c).fill = _fill("1A56DB")

mg(ws1, 6, 2, 8, 4)
c = ws1.cell(6, 2, "日報自動入力アプリ")
c.font = Font(name="Meiryo", bold=True, size=22, color="FFFFFF")
c.alignment = _align("center", "center")

mg(ws1, 9, 2, 10, 4)
c = ws1.cell(9, 2, "導入説明資料")
c.font = Font(name="Meiryo", bold=True, size=16, color="BFD7FF")
c.alignment = _align("center", "center")

mg(ws1, 12, 2, 13, 4)
c = ws1.cell(12, 2, "― 業務効率化に向けた月次作業報告書の自動化 ―")
c.font = Font(name="Meiryo", size=10, color="FFFFFF", italic=True)
c.alignment = _align("center", "center")

# 作成情報
info = [
    ("作成日",   TODAY),
    ("作成部署", ""),
    ("作成者",   ""),
    ("提出先",   ""),
]
for i, (label, val) in enumerate(info):
    r = 19 + i * 2
    rh(ws1, r, 20)
    mg(ws1, r, 2, r, 2)
    lc = ws1.cell(r, 2, label)
    lc.font = _font(bold=True, size=10, color="1A56DB")
    lc.fill = _fill("EFF6FF")
    lc.alignment = _align("center")
    lc.border = _border("BFD7FF")
    mg(ws1, r, 3, r, 4)
    vc = ws1.cell(r, 3, val)
    vc.font = _font(size=10)
    vc.fill = _fill("FFFFFF")
    vc.alignment = _align("left")
    vc.border = _border("BFD7FF")

mg(ws1, 30, 2, 30, 4)
c = ws1.cell(30, 2, "本資料は社内業務効率化ツールの導入説明を目的として作成されました。")
c.font = _font(size=8, color="94A3B8", italic=True)
c.alignment = _align("center")


# ════════════════════════════════════════════════════
# Sheet 2：アプリ概要
# ════════════════════════════════════════════════════
ws2 = wb.create_sheet("1. アプリ概要")
col_widths(ws2, [2, 18, 50, 20, 2])
ws2.sheet_view.showGridLines = False

r = 1
rh(ws2, r, 8); r += 1

# ページタイトル
mg(ws2, r, 1, r, 5); rh(ws2, r, 30)
c = ws2.cell(r, 1, "1.  アプリ概要")
c.font = _font(bold=True, size=16, color="1A56DB")
c.alignment = _align("left", "center")
r += 1
mg(ws2, r, 1, r, 5); rh(ws2, r, 20)
c = ws2.cell(r, 1, "日報自動入力アプリ  導入説明資料")
c.font = _font(size=9, color="94A3B8", italic=True)
c.alignment = _align("left", "center")
r += 1; spacer(ws2, r, 10); r += 1

# 1-1 導入目的
section_heading(ws2, r, "1-1　導入目的", "1E40AF"); r += 1
for txt in [
    "本アプリは、毎月会社から配布されるExcel形式の作業報告書（日報）への勤務時間・備考欄の記入作業を自動化するWebアプリケーションです。",
    "現状、社員は毎月末から月初にかけて1ヶ月分の勤務時間（開始・終了・休憩）および備考をExcelへ手作業で入力する業務が発生しています。",
    "本ツールを導入することで、月初に最低限の設定を行うだけで1ヶ月分のExcelへの自動入力・保存が完結し、作業工数の大幅な削減が見込まれます。",
]:
    rh(ws2, r, 36)
    mg(ws2, r, 2, r, 4)
    c = ws2.cell(r, 2, txt)
    c.font = _font(size=10); c.fill = _fill("F8FAFF")
    c.alignment = _align(wrap=True); c.border = _border("DBEAFE")
    r += 1
spacer(ws2, r, 8); r += 1

# 1-2 期待される効果
section_heading(ws2, r, "1-2　期待される効果", "059669"); r += 1
effects = [
    ("工数削減",       "月次の手入力作業（約15〜30分/月）をほぼゼロに削減。年間換算で数時間分の業務時間を創出します。"),
    ("入力ミスの防止", "手動入力に伴う時間の誤入力・記入漏れを排除し、正確な作業報告書を自動生成します。"),
    ("操作の簡便性",   "ブラウザのみで動作し、インストール不要。プログラミング知識がなくても月初に数分で操作が完了します。"),
    ("設定の引き継ぎ", "曜日別の勤務時間設定がブラウザに自動保存されるため、翌月以降は確認・変更のみで対応できます。"),
]
for label, text in effects:
    rh(ws2, r, 38)
    lc = ws2.cell(r, 2, label)
    lc.font = _font(bold=True, size=10, color="065F46"); lc.fill = _fill("D1FAE5")
    lc.alignment = _align("center"); lc.border = _border("6EE7B7")
    mg(ws2, r, 3, r, 4)
    tc = ws2.cell(r, 3, text)
    tc.font = _font(size=10); tc.fill = _fill("F0FDF4")
    tc.alignment = _align(wrap=True); tc.border = _border("6EE7B7")
    r += 1
spacer(ws2, r, 8); r += 1

# 1-3 対象・動作環境
section_heading(ws2, r, "1-3　対象・動作環境", "7C3AED"); r += 1
specs = [
    ("対象ユーザー", "Excel作業報告書（.xlsx形式）を毎月提出している社員"),
    ("対象ファイル", "会社から毎月配布されるExcel形式の作業報告書（.xlsx）"),
    ("動作環境",     "ブラウザ（Chrome / Edge / Safari など）。インストール・設定作業不要"),
    ("アクセス方法", "社内ネットワーク上のURL、またはローカル環境でブラウザから直接アクセス"),
    ("認証",         "ユーザー名・パスワードによるBasic認証（不正アクセス防止）"),
]
for label, text in specs:
    rh(ws2, r, 28)
    lc = ws2.cell(r, 2, label)
    lc.font = _font(bold=True, size=10, color="5B21B6"); lc.fill = _fill("EDE9FE")
    lc.alignment = _align("center"); lc.border = _border("C4B5FD")
    mg(ws2, r, 3, r, 4)
    tc = ws2.cell(r, 3, text)
    tc.font = _font(size=10); tc.fill = _fill("FAFAFA")
    tc.alignment = _align(wrap=True); tc.border = _border("C4B5FD")
    r += 1


# ════════════════════════════════════════════════════
# Sheet 3：操作手順
# ════════════════════════════════════════════════════
ws3 = wb.create_sheet("2. 操作手順")
col_widths(ws3, [2, 14, 52, 20, 2])
ws3.sheet_view.showGridLines = False

r = 1; spacer(ws3, r, 8); r += 1

mg(ws3, r, 1, r, 5); rh(ws3, r, 30)
c = ws3.cell(r, 1, "2.  操作手順")
c.font = _font(bold=True, size=16, color="1A56DB"); c.alignment = _align("left", "center")
r += 1
mg(ws3, r, 1, r, 5); rh(ws3, r, 18)
c = ws3.cell(r, 1, "月初の操作は以下7ステップで完了します。所要時間の目安：5分以内")
c.font = _font(size=9, color="94A3B8", italic=True); c.alignment = _align("left", "center")
r += 1; spacer(ws3, r, 10); r += 1

steps = [
    ("STEP 1", "アプリにアクセスする",
     "ブラウザでアプリのURLを開き、ユーザー名・パスワードを入力してログインします。",
     "1E40AF", "DBEAFE", "EFF6FF"),
    ("STEP 2", "対象月を確認する",
     "画面上部の「年・月」が自動的に当月に設定されます。\n別の月を対象にする場合は、数字を直接変更してください。",
     "065F46", "D1FAE5", "F0FDF4"),
    ("STEP 3", "曜日別の勤務時間を設定する",
     "月曜〜金曜それぞれの開始時間・終了時間・休憩時間を入力します。\n設定内容はブラウザに自動保存されるため、翌月以降は変更がある場合のみ修正すれば問題ありません。",
     "92400E", "FEF3C7", "FFFBEB"),
    ("STEP 4", "例外日を設定する（任意）",
     "残業・早退・休日出勤など、曜日別設定と異なる勤務時間の日がある場合は「＋ 例外日を追加」から個別に設定します。\n土日・祝日の休日出勤にも対応しています。例外日の設定は、他の設定よりも最優先で適用されます。",
     "7C3AED", "EDE9FE", "FAF5FF"),
    ("STEP 5", "有給取得日を選択する（任意）",
     "有給休暇を取得する日がある場合は「有給を取得する日がある」にチェックを入れ、表示されるカレンダーから該当日を選択します。\n土日・祝日も選択可能です。",
     "BE185D", "FCE7F3", "FFF0F7"),
    ("STEP 6", "Excelファイルを選択する",
     "「Excelファイルを選択」エリアをクリックし、会社から配布された作業報告書（.xlsx）を選択します。\n誤ったファイルを選択した場合は、右側の「✕」ボタンでキャンセルして選び直せます。",
     "0F766E", "CCFBF1", "F0FDFA"),
    ("STEP 7", "入力完了・ダウンロード",
     "「入力完了・ダウンロード」ボタンをクリックすると、自動入力済みのExcelファイルがダウンロードされます。\nダウンロードしたファイルを開いて内容を確認の上、所定の保存場所に保存してください。",
     "1E40AF", "DBEAFE", "EFF6FF"),
]

for step, title, desc, fc, hc, bc in steps:
    # ステップ見出し行
    rh(ws3, r, 20)
    sc = ws3.cell(r, 2, step)
    sc.font = Font(name="Meiryo", bold=True, size=9, color="FFFFFF")
    sc.fill = _fill(fc); sc.alignment = _align("center"); sc.border = _border(fc)
    mg(ws3, r, 3, r, 4)
    tc = ws3.cell(r, 3, title)
    tc.font = Font(name="Meiryo", bold=True, size=11, color=fc)
    tc.fill = _fill(hc); tc.alignment = _align(); tc.border = _border(hc)
    ws3.cell(r, 4).fill = _fill(hc); ws3.cell(r, 4).border = _border(hc)
    r += 1
    # 説明文行
    lines = desc.count('\n') + 1
    rh(ws3, r, max(36, lines * 26))
    mg(ws3, r, 2, r, 4)
    dc = ws3.cell(r, 2, desc)
    dc.font = _font(size=10); dc.fill = _fill(bc)
    dc.alignment = _align(wrap=True); dc.border = _border("E5E7EB")
    r += 1; spacer(ws3, r, 5); r += 1


# ════════════════════════════════════════════════════
# Sheet 4：自動入力ルール
# ════════════════════════════════════════════════════
ws4 = wb.create_sheet("3. 自動入力ルール")
col_widths(ws4, [2, 20, 30, 30, 2])
ws4.sheet_view.showGridLines = False

r = 1; spacer(ws4, r, 8); r += 1

mg(ws4, r, 1, r, 5); rh(ws4, r, 30)
c = ws4.cell(r, 1, "3.  自動入力ルール")
c.font = _font(bold=True, size=16, color="1A56DB"); c.alignment = _align("left", "center")
r += 1
mg(ws4, r, 1, r, 5); rh(ws4, r, 18)
c = ws4.cell(r, 1, "日の種類に応じて、以下のルールに従い自動で各セルへ入力されます。")
c.font = _font(size=9, color="94A3B8", italic=True); c.alignment = _align("left", "center")
r += 1; spacer(ws4, r, 10); r += 1

# テーブルヘッダー
rh(ws4, r, 26)
for col, label in zip([2, 3, 4], ["日の種類", "開始・終了・休憩時間", "備考欄の記入内容"]):
    hc = ws4.cell(r, col, label)
    hc.font = _font(bold=True, size=10, color="FFFFFF")
    hc.fill = _fill("1A56DB"); hc.alignment = _align("center"); hc.border = _border("1A56DB")
r += 1

rules = [
    ("出勤日（平日）",
     "曜日別設定の開始・終了時間\n休憩時間：1時間00分",
     "在宅勤務（出勤日の記入内容は変更可能）",
     "FFFFFF", "EFF6FF"),
    ("例外日（残業・早退・休日出勤）",
     "例外日に設定した開始・終了時間\n休憩時間：1時間00分",
     "在宅勤務（例外設定が他のすべての設定より優先）",
     "FFFBEB", "FEF3C7"),
    ("有給取得日",
     "空欄",
     "私用により、休暇",
     "FFF0F7", "FCE7F3"),
    ("祝日",
     "空欄",
     "祝日",
     "FFF5F5", "FEE2E2"),
    ("土日",
     "空欄",
     "空欄",
     "F9FAFB", "F3F4F6"),
]

for kind, times, note, bg, hbg in rules:
    lines = max(times.count('\n') + 1, note.count('\n') + 1)
    rh(ws4, r, max(36, lines * 24))
    kc = ws4.cell(r, 2, kind)
    kc.font = _font(bold=True, size=10); kc.fill = _fill(hbg)
    kc.alignment = _align("center"); kc.border = _border("E5E7EB")
    tc = ws4.cell(r, 3, times)
    tc.font = _font(size=10); tc.fill = _fill(bg)
    tc.alignment = _align("center", "center", wrap=True); tc.border = _border("E5E7EB")
    nc = ws4.cell(r, 4, note)
    nc.font = _font(size=10); nc.fill = _fill(bg)
    nc.alignment = _align("center", "center", wrap=True); nc.border = _border("E5E7EB")
    r += 1

spacer(ws4, r, 12); r += 1
section_heading(ws4, r, "備考", "64748B"); r += 1
notes_text = [
    "・祝日情報は外部API（jpholiday）から自動取得します。インターネット接続がない環境では祝日が空欄になる場合があります。",
    "・例外日の設定は、有給・祝日・土日の区別に関わらず最優先で適用されます（休日出勤の記録に使用可能）。",
    "・Excelファイルのヘッダー行（開始時間・終了時間・休憩時間・備考）を自動検索するため、列の位置が変わっても正常に動作します。",
]
for note in notes_text:
    rh(ws4, r, 28)
    mg(ws4, r, 2, r, 4)
    nc = ws4.cell(r, 2, note)
    nc.font = _font(size=9, color="475569"); nc.fill = _fill("F8FAFC")
    nc.alignment = _align(wrap=True); nc.border = _border("E2E8F0")
    r += 1


# ════════════════════════════════════════════════════
# Sheet 5：よくある質問
# ════════════════════════════════════════════════════
ws5 = wb.create_sheet("4. よくある質問")
col_widths(ws5, [2, 6, 60, 6, 2])
ws5.sheet_view.showGridLines = False

r = 1; spacer(ws5, r, 8); r += 1

mg(ws5, r, 1, r, 5); rh(ws5, r, 30)
c = ws5.cell(r, 1, "4.  よくある質問（Q&A）")
c.font = _font(bold=True, size=16, color="1A56DB"); c.alignment = _align("left", "center")
r += 1; spacer(ws5, r, 10); r += 1

faqs = [
    ("設定は毎月入力し直す必要がありますか？",
     "いいえ、不要です。曜日別の勤務時間・各種ラベル設定はブラウザに自動保存されます。次回起動時には前回の設定が自動的に読み込まれるため、変更がない場合はそのままご利用いただけます。"),
    ("休日出勤が発生した場合はどのように入力しますか？",
     "「＋ 例外日を追加」ボタンから対象日を選択し（土曜・日曜・祝日も選択可）、出勤時間を入力してください。例外日の設定は曜日別設定・有給・祝日より最優先で適用されます。"),
    ("有給を土日や祝日に取得した場合はどうなりますか？",
     "有給取得日カレンダーから土日・祝日も選択可能です。選択した日は備考欄に「私用により、休暇」が記入されます。"),
    ("Excelのファイルを誤って選択した場合はどうすればよいですか？",
     "ファイル名の右側に表示される「✕」ボタンをクリックすると選択がキャンセルされます。その後、再度クリックして正しいファイルを選び直してください。"),
    ("Excelファイルの列構成が異なる場合でも使用できますか？",
     "はい、使用できます。アプリは「開始時間・終了時間・休憩時間・備考」のヘッダー文字列を自動検索して対象列を特定するため、列の位置が変更された場合でも正常に動作します。ただし、ヘッダーの文字列が大きく異なる場合は、アプリ内のラベル設定から調整が必要です。"),
    ("アプリはどこで動作しますか？インストールが必要ですか？",
     "ブラウザのみで動作します。インストール・追加ソフトウェアは一切不要です。アプリが起動しているサーバー（PC）にアクセスできるネットワーク環境であれば、別のPCからも利用可能です。"),
]

for i, (q, a) in enumerate(faqs):
    # Q行
    rh(ws5, r, 22)
    mg(ws5, r, 2, r, 4)
    qc = ws5.cell(r, 2, f"Q{i+1}　{q}")
    qc.font = _font(bold=True, size=10, color="1E40AF")
    qc.fill = _fill("DBEAFE"); qc.alignment = _align(); qc.border = _border("BFDBFE")
    r += 1
    # A行
    lines = a.count('。') + 1
    rh(ws5, r, max(40, lines * 22))
    mg(ws5, r, 2, r, 4)
    ac = ws5.cell(r, 2, f"A.　{a}")
    ac.font = _font(size=10); ac.fill = _fill("FFFFFF")
    ac.alignment = _align(wrap=True); ac.border = _border("E5E7EB")
    r += 1; spacer(ws5, r, 6); r += 1


# ════════════════════════════════════════════════════
# 全シートのフッター・印刷設定
# ════════════════════════════════════════════════════
for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.oddFooter.center.text = f"日報自動入力アプリ 導入説明資料　{TODAY}　作成"
    ws.oddFooter.center.size = 8
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

OUTPATH = "/workspaces/DailyReportAutoFillApp/日報アプリ_導入説明資料.xlsx"
wb.save(OUTPATH)
print(f"作成完了: {OUTPATH}")
